from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from decimal import Decimal
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db import transaction, IntegrityError

import qrcode
import base64
from io import BytesIO
from django.urls import reverse
from django.views.decorators.http import require_POST

from stock.models import Article, DemandeMateriel, LigneDemande, Magasin
from django.contrib.auth.models import User
from accounts.permissions import verifier_permission

from core.models import Service

from stock.models import Fournisseur
from .models import TechnicienPrestataire

from .models import CampagneInventairePatrimoine, LigneInventairePatrimoine
from django.db.models import ProtectedError

from .models import (
    Immobilisation, TypeEquipement, CategoriePatrimoine,
    MouvementPatrimoine, ContratMaintenance, Intervention,
    ComptePrestataire, ImportPatrimoine, ParametresPatrimoine,
    Batiment, Etage, Bureau, Marque, Modele,
    TypeContrat, TechnicienPrestataire
)

# ══════════════════════════════════════════════════════════════════
# DÉCORATEUR PATRIMOINE
# ══════════════════════════════════════════════════════════════════

def patrimoine_required(view_func):
    """Vérifie que l'utilisateur a au moins une permission Patrimoine."""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/auth/login/')
        # Liste des permissions patrimoine existantes dans MenuAccess
        perms_pat = [
            'accounts.menu_pat_tickets', 'accounts.menu_pat_tech', 'accounts.menu_pat_dispatch',
            'accounts.menu_pat_historique', 'accounts.menu_pat_registre', 'accounts.menu_pat_sas',
            'accounts.menu_pat_contrats', 'accounts.menu_pat_import', 'accounts.menu_pat_inventaire',
            'accounts.menu_pat_rebuts', 'accounts.menu_pat_pertes', 'accounts.menu_pat_parametres'
        ]
        has_any = any(request.user.has_perm(p) for p in perms_pat)
        if not (request.user.is_staff or request.user.is_superuser or has_any):
            messages.error(request, "⛔ Accès non autorisé au module Patrimoine.")
            return redirect('/')
        return view_func(request, *args, **kwargs)
    return wrapper


# ══════════════════════════════════════════════════════════════════
# 1. REGISTRE GÉNÉRAL
# ══════════════════════════════════════════════════════════════════

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_registre')
def registre(request):
    qs = Immobilisation.objects.select_related(
        'type_equipement__categorie', 'bureau__etage__batiment',
        'service_affectation', 'marque', 'modele',
    ).exclude(statut='EN_ATTENTE').order_by('-date_creation')

    q = request.GET.get('q', '').strip()
    if '/scan/' in q:
        q = q.rstrip('/').split('/')[-1]
        
    categorie_id = request.GET.get('categorie', '')
    type_id      = request.GET.get('type', '')
    statut       = request.GET.get('statut', '')
    service_id   = request.GET.get('service', '')
    batiment_id  = request.GET.get('batiment', '')
    action       = request.GET.get('action', '')

    if q:
        qs = qs.filter(
            Q(code_patrimoine__icontains=q) | Q(numero_serie__icontains=q) |
            Q(nom_affichage__icontains=q)   | Q(marque__nom__icontains=q) |
            Q(modele__nom__icontains=q)     | Q(specs_techniques__icontains=q)
        ).distinct()
    if categorie_id: qs = qs.filter(type_equipement__categorie_id=categorie_id)
    if type_id:      qs = qs.filter(type_equipement_id=type_id)
    if statut:       qs = qs.filter(statut=statut)
    if service_id:   qs = qs.filter(service_affectation_id=service_id)
    if batiment_id:  qs = qs.filter(bureau__etage__batiment_id=batiment_id)
    if action:       qs = qs.filter(action_requise=action)

    stats = {
        'total':       qs.count(),
        'actifs':      qs.filter(statut='ACTIF').count(),
        'en_panne':    qs.filter(statut='EN_PANNE').count(),
        'en_attente':  Immobilisation.objects.filter(statut='EN_ATTENTE').count(),
        'val_totale':  qs.aggregate(t=Sum('valeur_acquisition'))['t'] or 0,
    }

    per_page = request.GET.get('per_page', '20')
    limite   = qs.count() if per_page == 'all' else int(per_page) if str(per_page).isdigit() else 20
    page     = Paginator(qs, limite).get_page(request.GET.get('page'))

    context = {
        'immos': page, 'today': timezone.now().date(), 'stats': stats,
        'categories': CategoriePatrimoine.objects.filter(est_active=True),
        'types': TypeEquipement.objects.filter(est_actif=True).select_related('categorie'),
        'batiments': Batiment.objects.all().order_by('code'),
        'q': q, 'categorie_id': categorie_id, 'type_id': type_id, 'statut': statut,
        'service_id': service_id, 'batiment_id': batiment_id, 'action': action, 'per_page': per_page,
        'statut_choices': Immobilisation.STATUT_CHOICES, 'action_choices': Immobilisation.ACTION_CHOICES,
    }
    return render(request, 'patrimoine/registre.html', context)


# ══════════════════════════════════════════════════════════════════
# 2. SAS & IMMATRICULATION
# ══════════════════════════════════════════════════════════════════

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_sas')
def sas(request):
    qs = Immobilisation.objects.filter(statut='EN_ATTENTE').order_by('-date_creation')
    return render(request, 'patrimoine/sas.html', {'immos': Paginator(qs, 25).get_page(request.GET.get('page')), 'nb_sas': qs.count()})

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_sas')
def valider_sas(request, pk):
    immo  = get_object_or_404(Immobilisation, pk=pk, statut='EN_ATTENTE')
    types = TypeEquipement.objects.filter(est_actif=True).select_related('categorie')

    if request.method == 'POST':
        try:
            te = TypeEquipement.objects.select_related('categorie').get(pk=request.POST.get('type_equipement'))
            immo.numero_serie         = request.POST.get('numero_serie') or ''
            immo.nom_affichage        = request.POST.get('nom_affichage') or ''
            immo.type_equipement_id   = te.pk
            immo.marque_id            = request.POST.get('marque') or None
            immo.modele_id            = request.POST.get('modele') or None
            immo.bureau_id            = request.POST.get('bureau') or None
            immo.service_affectation_id = request.POST.get('service') or None
            immo.emplacement_exact    = request.POST.get('emplacement_exact') or ''
            immo.date_mise_en_service = request.POST.get('date_mise_en_service') or None
            immo.garantie_expiration  = request.POST.get('garantie_expiration') or None
            immo.action_requise       = request.POST.get('action_requise', 'RAS')
            immo.notes                = request.POST.get('notes', '')

            if request.POST.get('valeur_acquisition'):
                immo.valeur_acquisition = Decimal(request.POST.get('valeur_acquisition'))
                immo.prix_depuis_stock  = False

            if request.POST.get('duree_amortissement_ans'):
                immo.duree_amortissement_ans = int(request.POST.get('duree_amortissement_ans'))

            specs = {}
            for champ in te.specs_schema:
                key = champ['key']
                val = request.POST.get(f'spec_{key}', '')
                if val: specs[key] = val
            immo.specs_techniques = specs

            immo.statut = 'ACTIF'
            immo.modifie_par = request.user

            annee = timezone.now().strftime('%y') 
            cat_code = te.categorie.code[:3].upper() 
            prefix = f"{annee}-{cat_code}-"

            while True:
                dernier_immo = Immobilisation.objects.filter(code_patrimoine__startswith=prefix).order_by('-code_patrimoine').first()
                nouveau_num = (int(dernier_immo.code_patrimoine.split('-')[-1]) + 1) if dernier_immo and dernier_immo.code_patrimoine else 1
                immo.code_patrimoine = f"{prefix}{nouveau_num:05d}"
                try:
                    with transaction.atomic():
                        immo.save()
                        MouvementPatrimoine.objects.create(
                            immobilisation=immo, type_mouvement='AFFECTATION', bureau_arrivee=immo.bureau,
                            service_arrivee=immo.service_affectation, date_mouvement=timezone.now().date(),
                            motif="Immatriculation initiale", effectue_par=request.user,
                        )
                    break 
                except IntegrityError:
                    continue

            messages.success(request, f"✅ Bien immatriculé sous le code {immo.code_patrimoine}.")
            return redirect('patrimoine_detail', pk=immo.pk)
        except Exception as e:
            messages.error(request, f"❌ Erreur : {e}")

    return render(request, 'patrimoine/valider_sas.html', {
        'immo': immo, 'types': types, 'marques': Marque.objects.all().order_by('nom'),
        'bureaux': Bureau.objects.select_related('etage__batiment').all().order_by('etage__batiment__code','etage__nom','nom'),
        'services': Service.objects.all().order_by('nom'), 'action_choices': Immobilisation.ACTION_CHOICES,
        'batiments': Batiment.objects.all().order_by('nom'),
    })


# ══════════════════════════════════════════════════════════════════
# 3. FICHE DÉTAIL & MODIFICATION
# ══════════════════════════════════════════════════════════════════

@login_required(login_url='/auth/login/')
@patrimoine_required
def fiche_detail(request, pk):
    immo = get_object_or_404(
        Immobilisation.objects.select_related(
            'type_equipement__categorie', 'bureau__etage__batiment', 'service_affectation', 'marque', 'modele',
            'fournisseur', 'bon_sortie_origine', 'article_stock', 'contrat_maintenance__prestataire',
        ).prefetch_related('mouvements', 'interventions'), pk=pk
    )
    return render(request, 'patrimoine/fiche_detail.html', {
        'immo': immo, 'mouvements': immo.mouvements.order_by('-date_mouvement')[:10],
        'interventions': immo.interventions.order_by('-date_signalement')[:10],
        'specs_schema': immo.type_equipement.specs_schema if immo.type_equipement else [],
        'specs': immo.specs_techniques,
    })

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_registre')
def modifier_immo(request, pk):
    immo  = get_object_or_404(Immobilisation, pk=pk)
    if request.method == 'POST':
        try:
            immo.code_patrimoine        = request.POST.get('code_patrimoine') or immo.code_patrimoine
            immo.numero_serie           = request.POST.get('numero_serie', immo.numero_serie)
            immo.nom_affichage          = request.POST.get('nom_affichage', immo.nom_affichage)
            immo.type_equipement_id     = request.POST.get('type_equipement', immo.type_equipement_id)
            immo.marque_id              = request.POST.get('marque') or None
            immo.modele_id              = request.POST.get('modele') or None
            immo.bureau_id              = request.POST.get('bureau') or None
            immo.service_affectation_id = request.POST.get('service') or None
            immo.emplacement_exact      = request.POST.get('emplacement_exact', '')
            immo.garantie_expiration    = request.POST.get('garantie_expiration') or None
            immo.action_requise         = request.POST.get('action_requise', 'RAS')
            immo.statut                 = request.POST.get('statut', immo.statut)
            immo.notes                  = request.POST.get('notes', '')

            if request.POST.get('valeur_acquisition'):
                immo.valeur_acquisition = Decimal(request.POST.get('valeur_acquisition'))

            if immo.type_equipement_id:
                te = TypeEquipement.objects.get(pk=immo.type_equipement_id)
                specs = dict(immo.specs_techniques)
                for champ in te.specs_schema:
                    key = champ['key']
                    val = request.POST.get(f'spec_{key}', '')
                    if val: specs[key] = val
                immo.specs_techniques = specs

            immo.modifie_par = request.user
            immo.save()
            messages.success(request, "✅ Bien mis à jour.")
            return redirect('patrimoine_detail', pk=immo.pk)
        except Exception as e:
            messages.error(request, f"❌ Erreur : {e}")

    return render(request, 'patrimoine/modifier_immo.html', {
        'immo': immo, 
        'types': TypeEquipement.objects.filter(est_actif=True).select_related('categorie'),
        'marques': Marque.objects.all().order_by('nom'), 
        'bureaux': Bureau.objects.select_related('etage__batiment').all(),
        'services': Service.objects.all().order_by('nom'),
        'batiments': Batiment.objects.all().order_by('nom'),
    })


# ══════════════════════════════════════════════════════════════════
# 4. MOUVEMENTS & CONTRATS
# ══════════════════════════════════════════════════════════════════

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_registre')
def creer_mouvement(request, pk):
    immo = get_object_or_404(Immobilisation, pk=pk)
    if request.method == 'POST':
        type_mv = request.POST.get('type_mouvement')
        try:
            mv = MouvementPatrimoine.objects.create(
                immobilisation=immo, type_mouvement=type_mv, bureau_depart=immo.bureau, service_depart=immo.service_affectation,
                bureau_arrivee_id=request.POST.get('bureau_arrivee') or None, service_arrivee_id=request.POST.get('service_arrivee') or None,
                date_mouvement=request.POST.get('date_mouvement') or timezone.now().date(), motif=request.POST.get('motif', ''), effectue_par=request.user, cree_par=request.user,
            )
            if type_mv == 'MUTATION':
                immo.bureau_id = request.POST.get('bureau_arrivee') or None
                immo.service_affectation_id = request.POST.get('service_arrivee') or None
            elif type_mv == 'REPARATION':
                immo.statut = 'EN_PANNE'
            elif type_mv == 'RETOUR_REPARATION':
                immo.statut = 'ACTIF'
            elif type_mv == 'REFORME':
                immo.statut = 'REFORME'
            immo.modifie_par = request.user
            immo.save()
            messages.success(request, f"✅ Mouvement enregistré.")
        except Exception as e:
            messages.error(request, f"❌ Erreur : {e}")
        return redirect('patrimoine_detail', pk=pk)

    return render(request, 'patrimoine/mouvement.html', {'immo': immo, 'types': MouvementPatrimoine.TYPE_CHOICES, 'bureaux': Bureau.objects.select_related('etage__batiment').all()})

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_historique')
def liste_mouvements(request):
    qs = MouvementPatrimoine.objects.select_related('immobilisation', 'bureau_depart__etage__batiment', 'bureau_arrivee__etage__batiment', 'effectue_par').order_by('-date_mouvement')
    return render(request, 'patrimoine/mouvements.html', {'mouvements': Paginator(qs, 25).get_page(request.GET.get('page'))})

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_contrats')
def liste_contrats(request):
    if request.method == 'POST' and 'creer_contrat' in request.POST:
        try:
            ContratMaintenance.objects.create(
                reference=request.POST.get('reference'),
                prestataire_id=request.POST.get('prestataire'),
                type_contrat_id=request.POST.get('type_contrat'),
                date_debut=request.POST.get('date_debut'),
                date_fin=request.POST.get('date_fin'),
                cout_annuel=request.POST.get('cout_annuel', 0),
                description=request.POST.get('description', ''),
                statut='ACTIF',
                cree_par=request.user
            )
            messages.success(request, "✅ Nouveau contrat créé avec succès.")
        except Exception as e:
            messages.error(request, f"❌ Erreur lors de la création : {e}")
        return redirect('patrimoine_contrats')

    qs = ContratMaintenance.objects.select_related('prestataire', 'type_contrat').annotate(nb_equip=Count('equipements')).order_by('date_fin')
    
    return render(request, 'patrimoine/contrats.html', {
        'contrats': qs, 
        'fournisseurs': Fournisseur.objects.all().order_by('raison_sociale'),
        'types_contrat': TypeContrat.objects.all().order_by('nom'),
        'nb_expirant': qs.filter(date_fin__lte=timezone.now().date() + timezone.timedelta(days=30), date_fin__gte=timezone.now().date()).count(),
        'nb_expires': qs.filter(date_fin__lt=timezone.now().date()).count(),
    })

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_contrats')
def detail_contrat(request, pk):
    contrat = get_object_or_404(ContratMaintenance.objects.select_related('prestataire').prefetch_related('equipements', 'interventions'), pk=pk)
    if request.method == 'POST' and 'save_contrat' in request.POST:
        try:
            contrat.reference = request.POST.get('reference', contrat.reference)
            contrat.date_debut = request.POST.get('date_debut', contrat.date_debut)
            contrat.date_fin = request.POST.get('date_fin', contrat.date_fin)
            contrat.cout_annuel = request.POST.get('cout_annuel', contrat.cout_annuel)
            contrat.description = request.POST.get('description', contrat.description)
            contrat.modifie_par = request.user
            contrat.save()
            messages.success(request, "✅ Contrat mis à jour.")
        except Exception as e:
            messages.error(request, f"❌ {e}")
        return redirect('patrimoine_contrat_detail', pk=pk)
    return render(request, 'patrimoine/contrat_detail.html', {'contrat': contrat, 'interventions': contrat.interventions.order_by('-date_signalement')[:20]})

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_contrats')
def assigner_equipements_contrat(request, contrat_id):
    from django.db.models import Q
    contrat = get_object_or_404(ContratMaintenance, id=contrat_id)
    
    equipements_disponibles = Immobilisation.objects.filter(
        Q(contrat_maintenance__isnull=True) | 
        Q(contrat_maintenance__statut='EXPIRE') | 
        Q(contrat_maintenance=contrat)
    ).select_related('type_equipement__categorie', 'service_affectation', 'bureau__etage__batiment')

    batiment_id = request.GET.get('batiment')
    service_id = request.GET.get('service')
    type_id = request.GET.get('type')

    if batiment_id:
        equipements_disponibles = equipements_disponibles.filter(bureau__etage__batiment_id=batiment_id)
    if service_id:
        equipements_disponibles = equipements_disponibles.filter(service_affectation_id=service_id)
    if type_id:
        equipements_disponibles = equipements_disponibles.filter(type_equipement_id=type_id)

    if request.method == 'POST':
        equipements_coches = request.POST.getlist('equipements')
        Immobilisation.objects.filter(contrat_maintenance=contrat).update(contrat_maintenance=None)
        if equipements_coches:
            Immobilisation.objects.filter(id__in=equipements_coches).update(contrat_maintenance=contrat)
        messages.success(request, "✅ La couverture du contrat a été mise à jour avec succès.")
        return redirect('patrimoine_contrats')

    context = {
        'contrat': contrat,
        'equipements': equipements_disponibles,
        'batiments': Batiment.objects.all().order_by('nom'),
        'services': Service.objects.all().order_by('nom'),
        'types_equipement': TypeEquipement.objects.all().order_by('nom'),
        'batiment_id': batiment_id,
        'service_id': service_id,
        'type_id': type_id,
    }
    return render(request, 'patrimoine/assigner_equipements_contrat.html', context)


# ══════════════════════════════════════════════════════════════════
# 6. GESTION DES INTERVENTIONS (LES 3 NOUVEAUX MODES)
# ══════════════════════════════════════════════════════════════════

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_tech')
def liste_interventions(request):
    params = ParametresPatrimoine.get_parametres()
    mode_visibilite = params.mode_visibilite_interventions
    qs = Intervention.objects.all()

    if not request.user.is_superuser:
        est_chef = request.user.is_staff
        try:
            profil = request.user.profil
            if mode_visibilite == 'DIRECT':
                if not est_chef:
                    domaines_ids = profil.domaines_intervention.values_list('id', flat=True)
                    qs = qs.filter(immobilisation__type_equipement__categorie_id__in=domaines_ids)
            elif mode_visibilite == 'DISPATCH':
                if not est_chef:
                    qs = qs.filter(intervenant=request.user)
        except AttributeError:
            if not est_chef:
                qs = Intervention.objects.none()

    qs = qs.select_related('immobilisation__type_equipement__categorie', 'immobilisation__service_affectation', 'intervenant', 'contrat').order_by('-date_signalement')
    per_page = request.GET.get('per_page', '15')
    limite = qs.count() or 1 if per_page == 'all' else int(per_page) if per_page.isdigit() else 15

    return render(request, 'patrimoine/interventions.html', {
        'interventions': Paginator(qs, limite).get_page(request.GET.get('page')),
        'mode_visibilite': mode_visibilite,
    })


@login_required(login_url='/auth/login/')
@patrimoine_required
def detail_intervention(request, intervention_id):
    intervention = get_object_or_404(Intervention, id=intervention_id)
    articles_catalogue = Article.objects.all().order_by('designation')
    fournisseurs = Fournisseur.objects.all().order_by('raison_sociale')
    techniciens_externes = TechnicienPrestataire.objects.filter(est_actif=True)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'prendre_en_charge':
            intervention.statut = 'EN_COURS'
            intervention.intervenant = request.user
            if not intervention.date_debut_intervention:
                intervention.date_debut_intervention = timezone.now()
            messages.success(request, "🔧 Vous avez pris en charge cette intervention.")

        elif action == 'demander_pieces':
            params = ParametresPatrimoine.get_parametres()
            if not params.magasin_pieces:
                messages.error(request, "❌ Aucun magasin de pièces configuré.")
                return redirect('detail_intervention', intervention_id=intervention.id)

            article_ids = request.POST.getlist('articles[]')
            quantites = request.POST.getlist('quantites[]')

            if article_ids:
                service_dem = intervention.immobilisation.service_affectation or (request.user.profil.service if hasattr(request.user, 'profil') else None)
                numero_demande = f"REQ-MAINT-{intervention.id}-{timezone.now().strftime('%y%m%d%H%M')}"
                nouvelle_demande = DemandeMateriel.objects.create(
                    numero_demande=numero_demande, demandeur=request.user, service_demandeur=service_dem,
                    magasin_cible=params.magasin_pieces, statut='EN_ATTENTE',
                    commentaire=f"Pièces pour l'intervention #{intervention.id}"
                )
                for aid, qte in zip(article_ids, quantites):
                    if aid and qte and int(qte) > 0:
                        LigneDemande.objects.create(demande=nouvelle_demande, article_id=aid, quantite_demandee=int(qte))
                intervention.demandes_pieces.add(nouvelle_demande)
                intervention.statut = 'EN_ATTENTE_PIECES'
                messages.success(request, "📦 Demande envoyée au magasin.")

        elif action == 'receptionner_demande':
            demande_id = request.POST.get('demande_id')
            demande = get_object_or_404(DemandeMateriel, id=demande_id)
            demande.statut = 'LIVREE'
            demande.save()
            demandes_restantes = intervention.demandes_pieces.exclude(statut__in=['ANNULEE', 'REJETEE', 'LIVREE', 'RECEPTIONNEE'])
            if not demandes_restantes.exists() and intervention.statut == 'EN_ATTENTE_PIECES':
                intervention.statut = 'EN_COURS'
            messages.success(request, "📦 Pièces réceptionnées.")

        elif action == 'deleguer_prestataire':
            prestataire_id = request.POST.get('prestataire')
            tech_id = request.POST.get('technicien_appele')
            sortie = request.POST.get('necessite_sortie') == 'on'

            intervention.prestataire_concerne_id = prestataire_id if prestataire_id else None
            intervention.technicien_appele_id = tech_id if tech_id else None
            intervention.date_appel_prestataire = timezone.now()
            intervention.necessite_sortie_reparation = sortie
            intervention.statut = 'EN_COURS'
            
            if sortie:
                MouvementPatrimoine.objects.create(
                    immobilisation=intervention.immobilisation,
                    type_mouvement='REPARATION',
                    bureau_depart=intervention.immobilisation.bureau,
                    service_depart=intervention.immobilisation.service_affectation,
                    date_mouvement=timezone.now().date(),
                    motif=f"Sortie pour réparation externe (Ticket #{intervention.id})",
                    effectue_par=request.user
                )
                intervention.immobilisation.statut = 'EN_PANNE'
                intervention.immobilisation.save()

            messages.success(request, "🤝 Intervention déléguée au prestataire externe.")

        elif action == 'retour_prestataire':
            intervention.date_retour_reelle = timezone.now()
            intervention.date_fin_intervention = timezone.now()
            intervention.actions_effectuees = request.POST.get('rapport', '')
            
            if 'rapport_scan' in request.FILES:
                intervention.rapport_prestataire_scan = request.FILES['rapport_scan']
            if 'bon_signe' in request.FILES:
                intervention.bon_sortie_signe_scan = request.FILES['bon_signe']
            
            intervention.statut = 'EN_ATTENTE_VALIDATION'
            
            if intervention.necessite_sortie_reparation:
                MouvementPatrimoine.objects.create(
                    immobilisation=intervention.immobilisation,
                    type_mouvement='RETOUR_REPARATION',
                    bureau_arrivee=intervention.immobilisation.bureau,
                    service_arrivee=intervention.immobilisation.service_affectation,
                    date_mouvement=timezone.now().date(),
                    motif=f"Retour de réparation externe (Ticket #{intervention.id})",
                    effectue_par=request.user
                )
            messages.success(request, "✅ Retour enregistré, rapport joint et dossier transmis pour validation.")

        elif action == 'soumettre_devis':
            intervention.frais_hors_contrat = Decimal(request.POST.get('montant_devis', 0))
            intervention.motif_frais_hors_contrat = request.POST.get('motif_devis', '')
            intervention.statut = 'EN_ATTENTE_DEVIS'
            intervention.devis_accepte = None
            messages.success(request, "📝 Devis soumis. En attente de l'accord de l'administration.")

        elif action == 'valider_devis' and request.user.is_superuser:
            intervention.devis_accepte = True
            intervention.statut = 'EN_COURS'
            messages.success(request, "✅ Devis validé ! L'intervention se poursuit.")

        elif action == 'refuser_devis' and request.user.is_superuser:
            intervention.devis_accepte = False
            intervention.statut = 'EN_COURS'
            messages.warning(request, "❌ Devis refusé.")
            
        elif action == 'resoudre':
            demandes_inutiles = intervention.demandes_pieces.exclude(statut__in=['LIVREE', 'RECEPTIONNEE', 'ANNULEE', 'REJETEE'])
            for d in demandes_inutiles:
                d.statut = 'ANNULEE'
                d.motif_refus = "Annulation automatique (Intervention terminée)"
                d.save()
            intervention.statut = 'EN_ATTENTE_VALIDATION'
            intervention.actions_effectuees = request.POST.get('rapport', intervention.actions_effectuees)
            diagnostic = request.POST.get('diagnostic')
            if diagnostic:
                intervention.diagnostic = diagnostic
            intervention.date_fin_intervention = timezone.now()
            messages.success(request, "✅ Réparation terminée. En attente de validation.")

        elif action == 'valider' and request.user.is_superuser:
            intervention.statut = 'RESOLUE'
            intervention.valide_par = request.user
            intervention.date_validation = timezone.now()
            machine = intervention.immobilisation
            machine.statut = 'ACTIF'
            machine.action_requise = 'RAS'
            machine.save()
            messages.success(request, "🎉 Intervention clôturée !")

        intervention.save()
        return redirect('detail_intervention', intervention_id=intervention.id)

    return render(request, 'patrimoine/detail_intervention.html', {
        'intervention': intervention, 
        'articles_catalogue': articles_catalogue,
        'fournisseurs': fournisseurs,
        'techniciens_externes': techniciens_externes
    })

@login_required(login_url='/auth/login/')
def declarer_panne_pc(request):
    if request.method == 'POST':
        immo_id = request.POST.get('equipement')
        description = request.POST.get('description', '')
        photo = request.FILES.get('photo')
        equipement = get_object_or_404(Immobilisation, id=immo_id)
        
        Intervention.objects.create(
            immobilisation=equipement, type_intervention='CURATIVE', statut='NOUVELLE',
            description_probleme=description, photo=photo, cree_par=request.user
        )
        equipement.statut = 'EN_PANNE'
        equipement.save()
        messages.success(request, f"✅ Panne signalée pour {equipement.nom_affichage}.")
        return redirect('patrimoine_mes_tickets')
        
    equipements = Immobilisation.objects.exclude(statut__in=['EN_ATTENTE', 'REFORME']).select_related('service_affectation', 'bureau').order_by('nom_affichage')
    
    params = ParametresPatrimoine.get_parametres()
    if not request.user.is_superuser and hasattr(request.user, 'profil'):
        profil = request.user.profil
        if params.perimetre_declaration == 'SERVICE' and profil.service:
            equipements = equipements.filter(service_affectation=profil.service)
        elif params.perimetre_declaration == 'BUREAU' and profil.bureau:
            equipements = equipements.filter(bureau=profil.bureau)
            
    return render(request, 'patrimoine/declarer_panne_pc.html', {'equipements': equipements})


def signaler_panne(request, immo_id):
    equipement = get_object_or_404(Immobilisation, id=immo_id)
    if request.method == 'POST':
        description = request.POST.get('description', '')
        photo = request.FILES.get('photo')
        Intervention.objects.create(
            immobilisation=equipement, type_intervention='CURATIVE', statut='NOUVELLE',
            description_probleme=description, photo=photo
        )
        return render(request, 'patrimoine/confirmer_signalement_panne.html')
    return render(request, 'patrimoine/signaler_panne.html', {'equipement': equipement})


@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_tech')
def creer_intervention(request, immo_pk):
    immo = get_object_or_404(Immobilisation, pk=immo_pk)
    if request.method == 'POST':
        try:
            est_prest = hasattr(request.user, 'compte_prestataire')
            type_inter = request.POST.get('type_intervention', 'CURATIVE')
            inter = Intervention.objects.create(
                immobilisation=immo, contrat_id=request.POST.get('contrat') or None, type_intervention=type_inter,
                intervenant=request.user, est_prestataire_externe=est_prest, date_signalement=timezone.now(),
                date_debut_intervention=request.POST.get('date_debut') or None, date_fin_intervention=request.POST.get('date_fin') or None,
                description_probleme=request.POST.get('description_probleme', ''), actions_effectuees=request.POST.get('actions_effectuees', ''),
                cout_main_oeuvre=request.POST.get('cout_mo', 0) or 0, cout_pieces=request.POST.get('cout_pieces', 0) or 0,
                cout_deplacement=request.POST.get('cout_deplacement', 0) or 0, statut='EN_ATTENTE_VALIDATION' if est_prest else 'EN_COURS',
                cree_par=request.user,
            )
            if type_inter == 'CURATIVE':
                immo.statut = 'EN_PANNE'
                immo.save()
            messages.success(request, "✅ Problème signalé avec succès.")
            if 'mobile' in request.GET:
                return redirect('patrimoine_scan', code=immo.code_patrimoine)
            return redirect('patrimoine_detail', pk=immo_pk)
        except Exception as e:
            messages.error(request, f"❌ Erreur : {e}")

    contrats = ContratMaintenance.objects.filter(equipements=immo, statut='ACTIF')
    return render(request, 'patrimoine/creer_intervention.html', {'immo': immo, 'contrats': contrats, 'types': Intervention.TYPE_CHOICES})


@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_tech')
def valider_intervention(request, pk):
    inter = get_object_or_404(Intervention, pk=pk)
    if request.method == 'POST':
        statut = request.POST.get('statut')
        inter.statut = statut
        inter.valide_par = request.user
        inter.date_validation = timezone.now()
        inter.commentaire_validation = request.POST.get('commentaire', '')
        inter.modifie_par = request.user
        inter.save()
        if statut in ['TERMINEE', 'CLOTUREE', 'VALIDEE', 'RESOLUE']:
            machine = inter.immobilisation
            machine.statut = 'ACTIF'
            machine.action_requise = 'RAS'
            machine.save()
        messages.success(request, "✅ Intervention validée et équipement mis à jour.")
    return redirect('patrimoine_detail', pk=inter.immobilisation_id)


# ══════════════════════════════════════════════════════════════════
# 7. PORTAIL PRESTATAIRE & PARAMÈTRES
# ══════════════════════════════════════════════════════════════════

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_tech')
def portail_prestataire(request):
    try:
        compte = request.user.compte_prestataire
    except Exception:
        messages.error(request, "⛔ Accès non autorisé.")
        return redirect('/')
    if not compte.est_actif:
        messages.error(request, "Votre compte prestataire est désactivé.")
        return redirect('/')
    contrats = compte.contrats_autorises.filter(statut='ACTIF').prefetch_related('equipements')
    interventions_recentes = Intervention.objects.filter(contrat__in=contrats).order_by('-date_signalement')[:20]
    return render(request, 'patrimoine/portail_prestataire.html', {'compte': compte, 'contrats': contrats, 'interventions_recentes': interventions_recentes})


@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_parametres')
def parametres(request):
    params = ParametresPatrimoine.objects.first()
    
    if request.method == 'POST':
        action  = request.POST.get('action')
        item_id = request.POST.get('item_id')
        
        try:
            if action == 'toggle_chef':
                user_id = request.POST.get('user_id')
                utilisateur = User.objects.get(id=user_id)
                
                # 🔒 Protection : vérifier que le user a un profil
                if not hasattr(utilisateur, 'profil') or utilisateur.profil is None:
                    messages.error(request, "❌ Cet utilisateur n'a pas de profil configuré. Créez-le d'abord dans la gestion des utilisateurs.")
                    return redirect('patrimoine_parametres')
                
                # Sécurité : on ne peut promouvoir que des gens de SON service (ou superuser)
                if not request.user.is_superuser and hasattr(request.user, 'profil') and request.user.profil.service:
                    if not utilisateur.profil.service or utilisateur.profil.service != request.user.profil.service:
                        messages.error(request, "⛔ Vous ne pouvez promouvoir que des membres de votre propre service.")
                        return redirect('patrimoine_parametres')
                
                nouveau_statut = not utilisateur.profil.est_chef_service
                utilisateur.profil.est_chef_service = nouveau_statut
                utilisateur.profil.save()
                titre = "promu chef de service" if nouveau_statut else "retiré du rôle de chef"
                messages.success(request, f"✅ {utilisateur.get_full_name() or utilisateur.username} a été {titre}.")

            elif action == 'save_visibilite':
                nouveau_mode = request.POST.get('mode_visibilite')
                if nouveau_mode in ['GLOBAL', 'DIRECT', 'DISPATCH']:
                    params.mode_visibilite_interventions = nouveau_mode
                
                nouveau_perimetre = request.POST.get('perimetre_declaration')
                if nouveau_perimetre in ['LIBRE', 'SERVICE', 'BUREAU']:
                    params.perimetre_declaration = nouveau_perimetre
                    
                magasin_id = request.POST.get('magasin_pieces')
                if magasin_id:
                    params.magasin_pieces_id = magasin_id
                else:
                    params.magasin_pieces = None
                    
                params.validation_inventaire_active = (request.POST.get('validation_inventaire') == 'on')
                params.save()
                
                validateurs_ids = request.POST.getlist('validateurs_inventaire')
                params.validateurs_inventaire.set(validateurs_ids)
                
                messages.success(request, "✅ Configuration et workflows mis à jour.")

            elif action == 'save_technicien':
                user_id = request.POST.get('user_id')
                categories_ids = request.POST.getlist('categories') 
                utilisateur = User.objects.get(id=user_id)
                
                # 🔒 Protection profil
                if not hasattr(utilisateur, 'profil') or utilisateur.profil is None:
                    messages.error(request, "❌ Cet utilisateur n'a pas de profil configuré.")
                    return redirect('patrimoine_parametres')
                    
                utilisateur.profil.domaines_intervention.set(categories_ids)
                messages.success(request, f"✅ Compétences mises à jour pour {utilisateur.get_full_name() or utilisateur.username}.")

            elif action == 'save_batiment':
                data = {'nom': request.POST.get('nom', '').upper(), 'code': request.POST.get('code', '').upper()}
                services_ids = request.POST.getlist('services')
                if item_id: 
                    batiment = Batiment.objects.get(pk=item_id)
                    Batiment.objects.filter(pk=item_id).update(**data, modifie_par=request.user)
                else: 
                    batiment = Batiment.objects.create(**data, cree_par=request.user)
                batiment.services.set(services_ids)
                for b in Bureau.objects.filter(etage__batiment=batiment):
                    b.services.set(services_ids)
                messages.success(request, "Bâtiment et ses bureaux mis à jour.")

            elif action == 'save_etage':
                nom = request.POST.get('nom', '').upper()
                batiment_id = request.POST.get('batiment_id')
                services_ids = request.POST.getlist('services')
                if item_id: 
                    etage = Etage.objects.get(pk=item_id)
                    Etage.objects.filter(pk=item_id).update(nom=nom, modifie_par=request.user)
                else: 
                    etage = Etage.objects.create(nom=nom, batiment_id=batiment_id, cree_par=request.user)
                etage.services.set(services_ids)
                for b in Bureau.objects.filter(etage=etage):
                    b.services.set(services_ids)
                messages.success(request, "Étage et ses bureaux mis à jour.")

            elif action == 'save_bureau':
                nom = request.POST.get('nom', '').upper()
                etage_id = request.POST.get('etage_id')
                services_ids = request.POST.getlist('services') 
                if item_id: 
                    bureau = Bureau.objects.get(pk=item_id)
                    bureau.nom = nom
                    bureau.modifie_par = request.user
                    bureau.save()
                else: 
                    bureau = Bureau.objects.create(nom=nom, etage_id=etage_id, cree_par=request.user)
                bureau.services.set(services_ids)
                messages.success(request, "Bureau enregistré et services liés.")
                
            elif action == 'assigner_bureau':
                user_id = request.POST.get('user_id')
                bureau_id = request.POST.get('bureau_id')
                utilisateur = User.objects.get(id=user_id)
                
                # 🔒 Protection profil
                if not hasattr(utilisateur, 'profil') or utilisateur.profil is None:
                    messages.error(request, "❌ Cet utilisateur n'a pas de profil configuré.")
                    return redirect('patrimoine_parametres')
                    
                utilisateur.profil.bureau_id = bureau_id if bureau_id else None
                utilisateur.profil.save()
                messages.success(request, f"✅ Bureau assigné à {utilisateur.get_full_name() or utilisateur.username}.")

            elif action == 'save_categorie':
                data = {'nom': request.POST.get('nom', '').upper(), 'code': request.POST.get('code', '').upper(), 'icone': request.POST.get('icone', 'fas fa-box'), 'couleur': request.POST.get('couleur', '#1c5b96'), 'modifie_par': request.user}
                if item_id: CategoriePatrimoine.objects.filter(pk=item_id).update(**data)
                else: data['cree_par'] = request.user; CategoriePatrimoine.objects.create(**data)
                messages.success(request, "Catégorie enregistrée.")
            elif action == 'delete_categorie':
                CategoriePatrimoine.objects.get(pk=item_id).delete()
                messages.success(request, "Catégorie supprimée.")

            elif action == 'save_type':
                data = {'categorie_id': request.POST.get('categorie'), 'nom': request.POST.get('nom', '').upper(), 'code': request.POST.get('code', '').upper(), 'duree_amortissement_defaut': int(request.POST.get('duree', 5)), 'mode_amortissement': request.POST.get('mode', 'LINEAIRE'), 'modifie_par': request.user}
                if item_id: TypeEquipement.objects.filter(pk=item_id).update(**data)
                else: data['cree_par'] = request.user; TypeEquipement.objects.create(**data)
                messages.success(request, "Type enregistré.")
            elif action == 'delete_type':
                TypeEquipement.objects.get(pk=item_id).delete()
                messages.success(request, "Type supprimé.")

            elif action == 'save_marque':
                nom = request.POST.get('nom', '').upper()
                if item_id: Marque.objects.filter(pk=item_id).update(nom=nom, modifie_par=request.user)
                else: Marque.objects.get_or_create(nom=nom, defaults={'cree_par': request.user})
                messages.success(request, "Marque enregistrée.")
            elif action == 'delete_marque':
                Marque.objects.get(pk=item_id).delete()
                messages.success(request, "Marque supprimée.")

            elif action == 'save_fournisseur':
                if item_id:
                    Fournisseur.objects.filter(id=item_id).update(raison_sociale=request.POST.get('raison_sociale'), telephone=request.POST.get('telephone'))
                    messages.success(request, "🏢 Fournisseur modifié.")
                else:
                    Fournisseur.objects.create(raison_sociale=request.POST.get('raison_sociale'), telephone=request.POST.get('telephone'), cree_par=request.user)
                    messages.success(request, "🏢 Nouveau fournisseur ajouté.")
            elif action == 'delete_fournisseur':
                Fournisseur.objects.filter(id=item_id).delete()
                messages.success(request, "🗑️ Fournisseur supprimé.")

            elif action == 'save_tech_externe':
                if item_id:
                    TechnicienPrestataire.objects.filter(id=item_id).update(fournisseur_id=request.POST.get('fournisseur_id'), nom=request.POST.get('nom'), telephone=request.POST.get('telephone'), specialite=request.POST.get('specialite', ''))
                    messages.success(request, "👷 Technicien externe mis à jour.")
                else:
                    TechnicienPrestataire.objects.create(fournisseur_id=request.POST.get('fournisseur_id'), nom=request.POST.get('nom'), telephone=request.POST.get('telephone'), specialite=request.POST.get('specialite', ''))
                    messages.success(request, "👷 Technicien externe ajouté.")
            elif action == 'delete_tech_externe':
                TechnicienPrestataire.objects.filter(id=item_id).delete()
                messages.success(request, "🗑️ Technicien externe supprimé.")

            elif action == 'save_type_contrat':
                if item_id:
                    TypeContrat.objects.filter(id=item_id).update(nom=request.POST.get('nom'), description=request.POST.get('description', ''))
                    messages.success(request, "📄 Type de contrat modifié.")
                else:
                    TypeContrat.objects.create(nom=request.POST.get('nom'), description=request.POST.get('description', ''))
                    messages.success(request, "📄 Nouveau type de contrat créé.")
            elif action == 'delete_type_contrat':
                TypeContrat.objects.filter(id=item_id).delete()
                messages.success(request, "🗑️ Type de contrat supprimé.")
                
        except ProtectedError:
            messages.error(request, "⛔ Impossible de supprimer — élément utilisé ailleurs.")
        except Exception as e:
            messages.error(request, f"❌ Erreur : {e}")

        return redirect('patrimoine_parametres')

    # 🔒 FILTRAGE : on n'affiche que les users AYANT un profil
    utilisateurs = User.objects.select_related(
        'profil', 'profil__service', 'profil__bureau'
    ).prefetch_related(
        'profil__domaines_intervention'
    ).filter(
        is_active=True, 
        profil__isnull=False
    ).order_by('first_name')
    
    from stock.models import Magasin, Fournisseur

    context = {
        'categories': CategoriePatrimoine.objects.prefetch_related('types_equipements').order_by('ordre', 'nom'),
        'types': TypeEquipement.objects.select_related('categorie').order_by('categorie', 'nom'),
        'marques': Marque.objects.all().order_by('nom'),
        'batiments': Batiment.objects.prefetch_related('etages__bureaux').order_by('code'),
        'params': params, 
        'utilisateurs': utilisateurs, 
        'magasins': Magasin.objects.all().order_by('nom'),
        'services': Service.objects.all().order_by('nom'),
        'fournisseurs': Fournisseur.objects.all().order_by('raison_sociale'),
        'techniciens': TechnicienPrestataire.objects.all().select_related('fournisseur').order_by('fournisseur__raison_sociale', 'nom'),
        'types_contrat': TypeContrat.objects.all().order_by('nom'),
    }
    return render(request, 'patrimoine/parametres.html', context)


@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_parametres')
def editer_schema(request, pk):
    te = get_object_or_404(TypeEquipement, pk=pk)
    if request.method == 'POST':
        try:
            schema_raw = request.POST.get('specs_schema', '[]')
            schema = json.loads(schema_raw)
            for champ in schema:
                if 'key' not in champ or 'label' not in champ:
                    raise ValueError(f"Champ invalide : {champ}")
            te.specs_schema = schema
            te.modifie_par = request.user
            te.save()
            messages.success(request, f"✅ Schéma mis à jour.")
            return redirect('patrimoine_parametres')
        except json.JSONDecodeError:
            messages.error(request, "❌ JSON invalide.")
        except Exception as e:
            messages.error(request, f"❌ Erreur : {e}")
    return render(request, 'patrimoine/editer_schema.html', {'te': te})


# ══════════════════════════════════════════════════════════════════
# 10. AJAX & UTILS
# ══════════════════════════════════════════════════════════════════

def ajax_modeles(request):
    marque_id = request.GET.get('marque')
    return JsonResponse(list(Modele.objects.filter(marque_id=marque_id).values('id', 'nom').order_by('nom')), safe=False)

def ajax_batiments(request):
    service_id = request.GET.get('service')
    qs = Batiment.objects.all()
    if service_id:
        qs = qs.filter(etages__bureaux__services__id=service_id).distinct()
    return JsonResponse(list(qs.values('id', 'nom').order_by('nom')), safe=False)

def ajax_localisation(request):
    service_id = request.GET.get('service')
    batiment_id = request.GET.get('batiment')
    etage_id = request.GET.get('etage')
    bureau_id = request.GET.get('bureau')

    qs_b = Bureau.objects.all()
    qs_e = Etage.objects.all()
    qs_bat = Batiment.objects.all()
    qs_s = Service.objects.all()

    if service_id:
        qs_b = qs_b.filter(services__id=service_id)
        qs_e = qs_e.filter(bureaux__services__id=service_id).distinct()
        qs_bat = qs_bat.filter(etages__bureaux__services__id=service_id).distinct()
    
    if batiment_id:
        qs_b = qs_b.filter(etage__batiment_id=batiment_id)
        qs_e = qs_e.filter(batiment_id=batiment_id)
        qs_s = qs_s.filter(bureaux_occupes__etage__batiment_id=batiment_id).distinct()
        
    if etage_id:
        qs_b = qs_b.filter(etage_id=etage_id)
        qs_bat = qs_bat.filter(etages__id=etage_id).distinct()
        qs_s = qs_s.filter(bureaux_occupes__etage_id=etage_id).distinct()
        
    if bureau_id:
        qs_e = qs_e.filter(bureaux__id=bureau_id).distinct()
        qs_bat = qs_bat.filter(etages__bureaux__id=bureau_id).distinct()
        qs_s = qs_s.filter(bureaux_occupes__id=bureau_id).distinct()

    return JsonResponse({
        'services': list(qs_s.values('id', 'nom').order_by('nom')),
        'batiments': list(qs_bat.values('id', 'nom').order_by('nom')),
        'etages': list(qs_e.values('id', 'nom').order_by('nom')),
        'bureaux': list(qs_b.values('id', 'nom').order_by('nom'))
    })

def ajax_specs_schema(request):
    try:
        te = TypeEquipement.objects.get(pk=request.GET.get('type'))
        return JsonResponse({'schema': te.specs_schema, 'duree': te.duree_amortissement_defaut, 'mode': te.mode_amortissement})
    except TypeEquipement.DoesNotExist:
        return JsonResponse({'schema': [], 'duree': 5, 'mode': 'LINEAIRE'})

def ajax_vnc(request, pk):
    try:
        immo = Immobilisation.objects.get(pk=pk)
        return JsonResponse({'vnc': str(immo.vnc), 'taux_amorti': str(immo.taux_amorti_pct), 'annees_ecoulees': str(immo.annees_ecoulees), 'est_totalement_amorti': immo.est_totalement_amorti})
    except Immobilisation.DoesNotExist:
        return JsonResponse({'error': 'not found'}, status=404)

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_registre')
def quick_edit(request, pk):
    if request.method != 'POST': return JsonResponse({'error': 'POST only'}, status=405)
    immo = get_object_or_404(Immobilisation, pk=pk)
    champ = request.POST.get('champ')
    val = request.POST.get('valeur')
    if champ not in {'statut', 'action_requise'}: return JsonResponse({'error': 'Champ non autorisé'}, status=400)
    setattr(immo, champ, val)
    immo.modifie_par = request.user
    immo.save(update_fields=[champ, 'modifie_par', 'date_modification'])
    return JsonResponse({'success': True, 'valeur': val, 'modifie_par': request.user.get_full_name() or request.user.username})

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_registre')
def etiquette_qr(request, pk):
    immo = get_object_or_404(Immobilisation, pk=pk)
    donnees_qr = request.build_absolute_uri(reverse('patrimoine_scan', args=[immo.code_patrimoine]))
    qr = qrcode.QRCode(version=1, box_size=10, border=1)
    qr.add_data(donnees_qr)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return render(request, 'patrimoine/etiquette_qr.html', {'immo': immo, 'qr_code': base64.b64encode(buffer.getvalue()).decode("utf-8")})

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_parametres')
def creer_type_equipement(request):
    if request.method == 'POST':
        try:
            specs_schema = json.loads(request.POST.get('specs_schema_cache', '[]'))
            TypeEquipement.objects.create(
                nom=request.POST.get('nom'), code=request.POST.get('code'), categorie_id=request.POST.get('categorie'),
                duree_amortissement_defaut=request.POST.get('duree_amortissement_defaut', 5), specs_schema=specs_schema, cree_par=request.user
            )
            messages.success(request, "Type créé avec succès !")
            return redirect('patrimoine_registre')
        except Exception as e:
            messages.error(request, f"Erreur : {e}")
    return render(request, 'patrimoine/creer_type_equipement.html', {'categories': CategoriePatrimoine.objects.all()})

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_parametres')
def api_type_specs(request, type_id):
    try:
        return JsonResponse({'specs': TypeEquipement.objects.get(id=type_id).specs_schema})
    except TypeEquipement.DoesNotExist:
        return JsonResponse({'specs': []})

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_sas')
@require_POST
def eclater_bien_sas(request):
    immo_id = request.POST.get('immo_id')
    noms_composants = request.POST.getlist('noms_composants[]')
    nombre = len(noms_composants) if noms_composants else int(request.POST.get('nombre', 2))
    if nombre < 2: return JsonResponse({'success': False, 'error': "Le nombre doit être au moins 2."})
    immo = get_object_or_404(Immobilisation, id=immo_id, statut='EN_ATTENTE')
    try:
        with transaction.atomic():
            nouvelle_valeur = (immo.valeur_acquisition or 0) / Decimal(nombre)
            nom_base = immo.nom_affichage or "Matériel"
            immo.nom_affichage = f"{nom_base} — {noms_composants[0]}" if noms_composants else f"{nom_base} (Partie 1)"
            immo.valeur_acquisition = nouvelle_valeur
            immo.save()
            for i in range(1, nombre):
                nom_clone = f"{nom_base} — {noms_composants[i]}" if noms_composants else f"{nom_base} (Partie {i+1})"
                Immobilisation.objects.create(
                    type_equipement=immo.type_equipement, article_stock=immo.article_stock, nom_affichage=nom_clone,
                    service_affectation=immo.service_affectation, bureau=immo.bureau, bon_sortie_origine=immo.bon_sortie_origine,
                    valeur_acquisition=nouvelle_valeur, statut='EN_ATTENTE', cree_par=request.user
                )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_sas')
def creer_immatriculation_directe(request):
    nouvelle_immo = Immobilisation.objects.create(nom_affichage="Nouveau Matériel (Saisie Directe)", statut='EN_ATTENTE', valeur_acquisition=0, cree_par=request.user)
    return redirect('patrimoine_valider_sas', pk=nouvelle_immo.id)

def scan_mobile(request, code):
    immo = get_object_or_404(Immobilisation, code_patrimoine=code)
    return render(request, 'patrimoine/scan_mobile.html', {'immo': immo})

# ══════════════════════════════════════════════════════════════════
# 11. EXPORTS ET IMPORTS EXCEL
# ══════════════════════════════════════════════════════════════════

COLONNES_FIXES = [
    'Service', 'Bâtiment', 'Étage', 'Bureau', 'Marque', 'Modèle',
    'N° de série', 'Code patrimoine / Asset Tag', 'Nom affichage',
    'Date acquisition (AAAA-MM-JJ)', 'Fournisseur', 'Valeur acquisition (FCFA)',
    'Garantie expiration (AAAA-MM-JJ)', 'Action requise', 'Notes',
]

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_registre')
def export_registre_excel(request):
    qs = Immobilisation.objects.select_related(
        'type_equipement__categorie', 'bureau__etage__batiment',
        'service_affectation', 'marque', 'modele',
    ).exclude(statut='EN_ATTENTE').order_by('-date_creation')

    q            = request.GET.get('q', '')
    categorie_id = request.GET.get('categorie', '')
    type_id      = request.GET.get('type', '')
    statut       = request.GET.get('statut', '')
    service_id   = request.GET.get('service', '')
    batiment_id  = request.GET.get('batiment', '')

    if q:
        qs = qs.filter(
            Q(code_patrimoine__icontains=q) | Q(numero_serie__icontains=q) |
            Q(nom_affichage__icontains=q)   | Q(marque__nom__icontains=q)
        ).distinct()
    if categorie_id: qs = qs.filter(type_equipement__categorie_id=categorie_id)
    if type_id:      qs = qs.filter(type_equipement_id=type_id)
    if statut:       qs = qs.filter(statut=statut)
    if service_id:   qs = qs.filter(service_affectation_id=service_id)
    if batiment_id:  qs = qs.filter(bureau__etage__batiment_id=batiment_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registre Patrimoine"

    hf = Font(bold=True, color='FFFFFF', size=11)
    hb = PatternFill('solid', fgColor='1C5B96')
    ca = Alignment(horizontal='center', vertical='center')

    headers = [
        'Code patrimoine', 'N° série', 'Désignation', 'Catégorie', 'Type',
        'Marque', 'Modèle', 'Statut', 'Service', 'Bâtiment', 'Bureau',
        'Date acquisition', 'Valeur acquisition (FCFA)', 'VNC (FCFA)', 'Taux amorti (%)',
        'Action requise', 'Créé par', 'Date création', 'Modifié par', 'Date modification',
    ]

    ws.row_dimensions[1].height = 35
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf; c.fill = hb; c.alignment = ca
        ws.column_dimensions[get_column_letter(i)].width = max(16, len(h)+4)

    for row_i, immo in enumerate(qs, 2):
        row = [
            immo.code_patrimoine or '', immo.numero_serie or '', immo.nom_affichage or '',
            immo.type_equipement.categorie.nom if immo.type_equipement_id else '',
            immo.type_equipement.nom if immo.type_equipement_id else '',
            immo.marque.nom if immo.marque_id else '', immo.modele.nom if immo.modele_id else '',
            immo.get_statut_display(),
            immo.service_affectation.nom if immo.service_affectation_id else '',
            str(immo.bureau.batiment.code) if immo.bureau_id else '',
            str(immo.bureau) if immo.bureau_id else '',
            str(immo.date_acquisition) if immo.date_acquisition else '',
            float(immo.valeur_acquisition), float(immo.vnc), float(immo.taux_amorti_pct),
            immo.get_action_requise_display(),
            immo.cree_par.get_full_name() if immo.cree_par_id else '',
            immo.date_creation.strftime('%d/%m/%Y %H:%M') if immo.date_creation else '',
            immo.modifie_par.get_full_name() if immo.modifie_par_id else '',
            immo.date_modification.strftime('%d/%m/%Y %H:%M') if immo.date_modification else '',
        ]
        for col_i, val in enumerate(row, 1):
            ws.cell(row=row_i, column=col_i, value=val)
        if row_i % 2 == 0:
            for col_i in range(1, len(headers)+1):
                ws.cell(row=row_i, column=col_i).fill = PatternFill('solid', fgColor='F0F6FF')

    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="patrimoine_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    wb.save(response)
    return response

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_import')
def telecharger_template(request, type_id):
    te = get_object_or_404(TypeEquipement, pk=type_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = te.nom[:30]

    header_font    = Font(bold=True, color='FFFFFF', size=11)
    header_fill_b  = PatternFill('solid', fgColor='1C5B96')   
    header_fill_p  = PatternFill('solid', fgColor='6F42C1')   
    center         = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin           = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.row_dimensions[1].height = 30
    ws.merge_cells(f'A1:{get_column_letter(len(COLONNES_FIXES) + len(te.specs_schema))}1')
    ws['A1'] = f"TEMPLATE IMPORT PATRIMOINE — {te.categorie.nom.upper()} / {te.nom.upper()}  |  Colonnes bleues = obligatoires/fixes  |  Colonnes violettes = specs techniques du type"
    ws['A1'].font = Font(bold=True, size=10, color='333333')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill('solid', fgColor='F0F3F7')

    ws.row_dimensions[2].height = 40
    for col_idx, nom_col in enumerate(COLONNES_FIXES, start=1):
        cell = ws.cell(row=2, column=col_idx, value=nom_col)
        cell.font = header_font; cell.fill = header_fill_b; cell.alignment = center; cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(nom_col) + 4)

    for i, spec in enumerate(te.specs_schema):
        col_idx = len(COLONNES_FIXES) + i + 1
        label = spec.get('label', spec.get('key', f'Spec{i+1}'))
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = header_font; cell.fill = header_fill_p; cell.alignment = center; cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(label) + 4)

    ws.row_dimensions[3].height = 20
    exemple = [
        'DIRECTION INFORMATIQUE', 'N', '1er Étage', 'BUREAU INFO', 'HP', 'HP DTP 300 G6 MT', 'SN123456789', 'CHU-INFO-2026-001',
        'UC HP DIRECTION INFO', '2024-01-15', 'BUDGET CHU ANGRÉ', '450000', '2027-01-15', 'RAS', '',
    ]
    for col_idx, val in enumerate(exemple, start=1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.font = Font(italic=True, color='888888', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.freeze_panes = 'A3'

    ws2 = wb.create_sheet("Guide")
    ws2['A1'] = "GUIDE D'UTILISATION"
    ws2['A1'].font = Font(bold=True, size=14)
    guide = [
        ("", ""), ("Colonne", "Description"),
        ("Service", "Nom exact du service (ex: DIRECTION INFORMATIQUE)"),
        ("Bâtiment", "Code ou nom du bâtiment (ex: N, A, BLOC TECHNIQUE)"),
        ("Étage", "Nom de l'étage (ex: RDC, 1er Étage, Sous-sol)"),
        ("Bureau", "Nom du bureau/salle"),
        ("N° de série", "Numéro de série physique de l'appareil"),
        ("Code patrimoine", "Asset Tag — laisser vide si non encore immatriculé"),
        ("Date acquisition", "Format AAAA-MM-JJ obligatoire pour les calculs d'amortissement"),
        ("Valeur acquisition", "Montant en FCFA — sans espace ni symbole"),
        ("Action requise", f"Valeurs possibles: {', '.join([c[0] for c in Immobilisation.ACTION_CHOICES])}"),
        ("", ""), ("Note", "La ligne 3 de chaque onglet est un exemple — ne pas supprimer, effacer les valeurs si besoin"),
    ]
    for i, (col, desc) in enumerate(guide, start=2):
        ws2.cell(row=i, column=1, value=col).font = Font(bold=True if i == 4 else False)
        ws2.cell(row=i, column=2, value=desc)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 65

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="template_{te.code}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_import')
def import_excel(request):
    types = TypeEquipement.objects.filter(est_actif=True).select_related('categorie')

    if request.method == 'POST':
        fichier   = request.FILES.get('fichier')
        type_id   = request.POST.get('type_equipement')

        if not fichier or not type_id:
            messages.error(request, "Veuillez choisir un type et un fichier.")
            return render(request, 'patrimoine/import.html', {'types': types})

        if not fichier.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Format accepté : .xlsx ou .xls uniquement.")
            return render(request, 'patrimoine/import.html', {'types': types})

        te = get_object_or_404(TypeEquipement, pk=type_id)

        try:
            wb = openpyxl.load_workbook(fichier, data_only=True)
            ws = wb.active

            header_row  = None
            header_map  = {}
            for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), start=1):
                row_vals = [str(c).strip() if c else '' for c in row]
                if any('service' in v.lower() for v in row_vals):
                    header_row = row_idx
                    for col_idx, val in enumerate(row_vals):
                        if val:
                            header_map[val.lower().strip()] = col_idx
                    break

            if header_row is None:
                messages.error(request, "En-têtes introuvables. Utilisez le template fourni.")
                return redirect('patrimoine_import')

            def get_val(row, key_fragments):
                for map_key, col_idx in header_map.items():
                    if any(frag.lower() in map_key for frag in key_fragments):
                        val = row[col_idx] if col_idx < len(row) else None
                        return str(val).strip() if val else ''
                return ''

            nb_crees = nb_maj = nb_err = 0
            log_err  = []

            for row_num, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
                if not any(row): continue
                if row_num == header_row + 1 and get_val(row, ['service']).upper() in ('SERVICE', 'DIRECTION INFORMATIQUE', 'NOM DU SERVICE'):
                    continue

                try:
                    from core.models import Service
                    from stock.models import Fournisseur

                    nom_bat = get_val(row, ['timent', 'bat'])
                    nom_eta = get_val(row, ['tage'])
                    nom_bur = get_val(row, ['bureau', 'salle'])
                    nom_svc = get_val(row, ['service'])

                    batiment = Batiment.objects.filter(Q(code__iexact=nom_bat) | Q(nom__iexact=nom_bat)).first()
                    if not batiment and nom_bat: batiment = Batiment.objects.create(code=nom_bat[:10].upper(), nom=nom_bat.upper(), cree_par=request.user)

                    etage = None
                    if batiment and nom_eta: etage, _ = Etage.objects.get_or_create(batiment=batiment, nom=nom_eta.upper(), defaults={'cree_par': request.user})

                    bureau = None
                    if etage and nom_bur: bureau, _ = Bureau.objects.get_or_create(etage=etage, nom=nom_bur.upper(), defaults={'cree_par': request.user})

                    service = Service.objects.filter(nom__iexact=nom_svc).first() if nom_svc else None

                    nom_marque = get_val(row, ['marque'])
                    nom_modele = get_val(row, ['mod'])
                    marque = Marque.objects.get_or_create(nom=nom_marque.upper(), defaults={'cree_par': request.user})[0] if nom_marque else None
                    modele = Modele.objects.get_or_create(marque=marque, nom=nom_modele.upper(), defaults={'cree_par': request.user})[0] if marque and nom_modele else None

                    code_pat  = get_val(row, ['asset', 'code', 'inventaire'])
                    num_serie = get_val(row, ['rie', 'sn'])
                    nom_aff   = get_val(row, ['nom', 'affich', 'equipement'])
                    date_acq_raw = get_val(row, ['acquisition', 'date'])
                    valeur_raw   = get_val(row, ['valeur', 'montant', 'fcfa'])
                    garantie_raw = get_val(row, ['garantie', 'expiration'])

                    date_acq = None
                    if date_acq_raw:
                        from datetime import datetime
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                            try: date_acq = datetime.strptime(str(date_acq_raw)[:10], fmt).date(); break
                            except ValueError: pass

                    valeur = Decimal('0.00')
                    if valeur_raw:
                        try: valeur = Decimal(str(valeur_raw).replace(' ', '').replace(',', '.'))
                        except Exception: pass

                    garantie = None
                    if garantie_raw:
                        from datetime import datetime
                        try: garantie = datetime.strptime(str(garantie_raw)[:10], '%Y-%m-%d').date()
                        except Exception: pass

                    action = get_val(row, ['action'])
                    if action not in [c[0] for c in Immobilisation.ACTION_CHOICES]: action = 'RAS'
                    notes  = get_val(row, ['note'])

                    specs = {}
                    for spec in te.specs_schema:
                        key = spec['key']; label = spec.get('label', key).lower()
                        val = get_val(row, [label, key])
                        if val: specs[key] = val

                    nom_fourn = get_val(row, ['fournisseur'])
                    fournisseur = Fournisseur.objects.filter(raison_sociale__icontains=nom_fourn).first() if nom_fourn else None

                    lookup = {}
                    if code_pat and code_pat.upper() not in ('NA', 'N/A', ''): lookup['code_patrimoine'] = code_pat
                    elif num_serie: lookup['numero_serie'] = num_serie
                    else: lookup = None

                    defaults = {
                        'type_equipement': te, 'nom_affichage': nom_aff or '', 'numero_serie': num_serie or '',
                        'marque': marque, 'modele': modele, 'bureau': bureau, 'service_affectation': service,
                        'date_acquisition': date_acq, 'valeur_acquisition': valeur, 'garantie_expiration': garantie,
                        'action_requise': action, 'notes': notes, 'specs_techniques': specs, 'fournisseur': fournisseur,
                        'statut': 'EN_ATTENTE' if not code_pat or code_pat.upper() in ('NA', 'N/A') else 'ACTIF',
                        'cree_par': request.user, 'reference_inventaire': fichier.name[:50],
                    }

                    if lookup:
                        obj, created = Immobilisation.objects.update_or_create(**lookup, defaults=defaults)
                        if created: nb_crees += 1
                        else: nb_maj += 1
                    else:
                        Immobilisation.objects.create(**defaults)
                        nb_crees += 1

                except Exception as e:
                    nb_err += 1
                    log_err.append({'ligne': row_num, 'erreur': str(e)})

            statut_log = 'OK' if nb_err == 0 else ('PARTIEL' if nb_crees + nb_maj > 0 else 'ECHEC')
            log = ImportPatrimoine.objects.create(
                type_equipement=te, nb_lignes_traitees=nb_crees + nb_maj + nb_err, nb_crees=nb_crees,
                nb_mis_a_jour=nb_maj, nb_erreurs=nb_err, log_erreurs=log_err, statut=statut_log, cree_par=request.user,
            )

            if nb_err == 0: messages.success(request, f"✅ Import réussi — {nb_crees} créés, {nb_maj} mis à jour.")
            else: messages.warning(request, f"⚠️ Import partiel — {nb_crees} créés, {nb_maj} mis à jour, {nb_err} erreurs.")

            return redirect('patrimoine_import_log', pk=log.pk)

        except Exception as e:
            messages.error(request, f"❌ Erreur lecture fichier : {e}")
            return redirect('patrimoine_import')

    context = {'types': types, 'logs_recents': ImportPatrimoine.objects.order_by('-date_creation')[:5]}
    return render(request, 'patrimoine/import.html', context)

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_import')
def detail_import_log(request, pk):
    log = get_object_or_404(ImportPatrimoine, pk=pk)
    return render(request, 'patrimoine/import_log.html', {'log': log})


# ══════════════════════════════════════════════════════════════════
# 12. ESPACES DÉDIÉS : TICKETS, DISPATCH ET KANBAN
# ══════════════════════════════════════════════════════════════════

@login_required(login_url='/auth/login/')
def mes_tickets(request):
    qs = Intervention.objects.filter(cree_par=request.user).select_related(
        'immobilisation__type_equipement__categorie',
        'immobilisation__service_affectation',
        'intervenant'
    ).order_by('-date_signalement')

    tab = request.GET.get('tab', 'encours')
    statuts_historique = ['RESOLUE', 'TERMINEE', 'CLOTUREE', 'ANNULE', 'ANNULEE', 'REJETEE', 'FERME', 'FERMEE']
    
    if tab == 'historique':
        qs = qs.filter(statut__in=statuts_historique)
    else:
        qs = qs.exclude(statut__in=statuts_historique)

    per_page = request.GET.get('per_page', '10')
    limite = qs.count() or 1 if per_page == 'all' else int(per_page) if str(per_page).isdigit() else 10
    page = Paginator(qs, limite).get_page(request.GET.get('page'))

    all_qs = Intervention.objects.filter(cree_par=request.user)
    stats = {
        'en_cours': all_qs.exclude(statut__in=statuts_historique).count(),
        'resolus': all_qs.filter(statut__in=statuts_historique).count(),
    }

    return render(request, 'patrimoine/mes_tickets.html', {
        'tickets': page, 'stats': stats, 'per_page': per_page, 'current_tab': tab
    })


@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_dispatch')
def dispatch_interventions(request):
    # ── Seuls les chefs de service peuvent dispatcher ──
    if not hasattr(request.user, 'profil') or not request.user.profil.est_chef_service:
        messages.error(request, "⛔ Accès réservé aux chefs de service.")
        return redirect('/')
    
    service_chef = request.user.profil.service
    if not service_chef:
        messages.error(request, "⛔ Aucun service assigné à votre profil.")
        return redirect('/')
        
    params = ParametresPatrimoine.get_parametres()
    mode_actuel = params.mode_visibilite_interventions

    if request.method == 'POST':
        intervention_id = request.POST.get('intervention_id')
        technicien_id = request.POST.get('technicien_id')
        intervention = get_object_or_404(Intervention, id=intervention_id)
        
        # Sécurité : on ne peut dispatcher que les pannes de SON service
        if intervention.immobilisation.service_affectation != service_chef:
            messages.error(request, "⛔ Cette panne ne concerne pas votre service.")
            return redirect('patrimoine_dispatch')
            
        if technicien_id:
            technicien = get_object_or_404(User, id=technicien_id)
            intervention.intervenant = technicien
            intervention.statut = 'PLANIFIEE'
            intervention.modifie_par = request.user
            intervention.save()
            messages.success(request, f"✅ Panne assignée à {technicien.get_full_name() or technicien.username}.")
        return redirect('patrimoine_dispatch')

    if mode_actuel == 'DISPATCH':
        pannes_en_attente = Intervention.objects.filter(
            statut='NOUVELLE', 
            intervenant__isnull=True,
            immobilisation__service_affectation=service_chef
        ).select_related(
            'immobilisation__type_equipement__categorie',
            'immobilisation__service_affectation',
            'cree_par'
        ).order_by('date_signalement')
    else:
        pannes_en_attente = Intervention.objects.none()

    # Techniciens rattachés au même service
    techniciens = User.objects.filter(
        is_active=True,
        profil__service=service_chef
    ).order_by('first_name')

    return render(request, 'patrimoine/dispatch.html', {
        'pannes_en_attente': pannes_en_attente,
        'techniciens': techniciens,
        'mode_actuel': mode_actuel,
        'service_chef': service_chef,
    })

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_tech')
def mes_interventions_tech(request):
    params = ParametresPatrimoine.get_parametres()
    mode_visibilite = params.mode_visibilite_interventions

    mes_pannes = Intervention.objects.filter(intervenant=request.user).select_related(
        'immobilisation__type_equipement__categorie',
        'immobilisation__service_affectation',
        'cree_par'
    )

    a_faire_perso = mes_pannes.filter(statut__in=['NOUVELLE', 'PLANIFIEE']).order_by('-date_signalement')
    en_cours = mes_pannes.filter(statut__in=['EN_COURS', 'EN_ATTENTE_PIECES']).order_by('-date_signalement')
    a_valider = mes_pannes.filter(statut='EN_ATTENTE_VALIDATION').order_by('-date_fin_intervention')
    historique_perso = mes_pannes.filter(statut__in=['RESOLUE', 'CLOTUREE', 'TERMINEE']).order_by('-date_fin_intervention')

    pannes_libres = Intervention.objects.none()
    if mode_visibilite != 'DISPATCH':
        base_libres = Intervention.objects.filter(
            statut='NOUVELLE', intervenant__isnull=True
        ).select_related(
            'immobilisation__type_equipement__categorie',
            'immobilisation__service_affectation',
            'cree_par'
        ).order_by('date_signalement')

        if mode_visibilite == 'GLOBAL':
            pannes_libres = base_libres
        elif mode_visibilite == 'DIRECT' and hasattr(request.user, 'profil'):
            domaines_ids = request.user.profil.domaines_intervention.values_list('id', flat=True)
            pannes_libres = base_libres.filter(immobilisation__type_equipement__categorie_id__in=domaines_ids)

    types_equipements = TypeEquipement.objects.all().order_by('nom')

    return render(request, 'patrimoine/mes_interventions.html', {
        'a_faire_perso': a_faire_perso,
        'pannes_libres': pannes_libres,
        'en_cours': en_cours,
        'a_valider': a_valider,
        'historique_perso': historique_perso,
        'mode_visibilite': mode_visibilite,
        'types_equipements': types_equipements,
    })


@login_required(login_url='/auth/login/')
def suivi_ticket(request, pk):
    intervention = get_object_or_404(Intervention, pk=pk)
    if intervention.cree_par != request.user and not request.user.is_superuser and not request.user.has_perm('accounts.menu_pat_tech'):
        messages.error(request, "⛔ Vous n'êtes pas autorisé à voir ce ticket.")
        return redirect('patrimoine_mes_tickets')
    return render(request, 'patrimoine/suivi_ticket.html', {'intervention': intervention})


@login_required(login_url='/auth/login/')
def declarer_panne_pc(request):
    if request.method == 'POST':
        immo_id = request.POST.get('equipement')
        description = request.POST.get('description', '')
        urgence = request.POST.get('degre_urgence', 'MOYENNE')
        photo = request.FILES.get('photo')
        equipement = get_object_or_404(Immobilisation, id=immo_id)
        
        Intervention.objects.create(
            immobilisation=equipement, type_intervention='CURATIVE', statut='NOUVELLE',
            description_probleme=description, degre_urgence=urgence, photo=photo, cree_par=request.user
        )
        equipement.statut = 'EN_PANNE'
        equipement.save()
        messages.success(request, f"✅ Panne signalée pour {equipement.nom_affichage}.")
        return redirect('patrimoine_mes_tickets')
        
    equipements = Immobilisation.objects.exclude(statut__in=['EN_ATTENTE', 'REFORME']).select_related('service_affectation', 'bureau').order_by('nom_affichage')
    
    try:
        params = ParametresPatrimoine.get_parametres()
        if not request.user.is_superuser and hasattr(request.user, 'profil'):
            profil = request.user.profil
            if params.perimetre_declaration == 'SERVICE' and profil.service:
                equipements = equipements.filter(service_affectation=profil.service)
            elif params.perimetre_declaration == 'BUREAU' and profil.bureau:
                equipements = equipements.filter(bureau=profil.bureau)
    except Exception:
        pass
            
    return render(request, 'patrimoine/declarer_panne_pc.html', {'equipements': equipements})


@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_tech')
def imprimer_bon_sortie_reparation(request, pk):
    intervention = get_object_or_404(Intervention, pk=pk)
    return render(request, 'patrimoine/bon_sortie_reparation_print.html', {'intervention': intervention})




@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_inventaire')
def patrimoine_campagnes_inventaire(request):
    campagnes = CampagneInventairePatrimoine.objects.all().select_related('categorie_cible', 'batiment_cible', 'responsable').order_by('-date_debut')
    categories = CategoriePatrimoine.objects.filter(est_active=True).order_by('nom')
    batiments = Batiment.objects.all().order_by('nom')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'save_campagne':
            item_id = request.POST.get('item_id')
            titre = request.POST.get('titre')
            reference = request.POST.get('reference')
            date_debut = request.POST.get('date_debut')
            date_fin_prevue = request.POST.get('date_fin_prevue') or None
            categorie_id = request.POST.get('categorie_cible') or None
            batiment_id = request.POST.get('batiment_cible') or None

            if item_id:
                CampagneInventairePatrimoine.objects.filter(id=item_id).update(
                    titre=titre, reference=reference, date_debut=date_debut,
                    date_fin_prevue=date_fin_prevue, categorie_cible_id=categorie_id,
                    batiment_cible_id=batiment_id, modifie_par=request.user
                )
                messages.success(request, "Campagne modifiée avec succès.")
            else:
                CampagneInventairePatrimoine.objects.create(
                    titre=titre, reference=reference, date_debut=date_debut,
                    date_fin_prevue=date_fin_prevue, categorie_cible_id=categorie_id,
                    batiment_cible_id=batiment_id, responsable=request.user,
                    cree_par=request.user, statut='BROUILLON'
                )
                messages.success(request, "Nouvelle campagne d'inventaire créée.")

        elif action == 'delete_campagne':
            item_id = request.POST.get('item_id')
            try:
                CampagneInventairePatrimoine.objects.filter(id=item_id).delete()
                messages.success(request, "Campagne supprimée.")
            except ProtectedError:
                messages.error(request, "⛔ Impossible de supprimer cette campagne car des scans ont déjà été effectués.")

        return redirect('patrimoine_campagnes_inventaire')

    context = {'campagnes': campagnes, 'categories': categories, 'batiments': batiments}
    return render(request, 'patrimoine/patrimoine_campagnes_inventaire.html', context)


# ==============================================================================
# 2. DÉTAIL DE LA CAMPAGNE (Génération, Lancement, Clôture, Stats)
# ==============================================================================
@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_inventaire')
def detail_campagne_inventaire(request, campagne_id):
    campagne = get_object_or_404(CampagneInventairePatrimoine, id=campagne_id)
    
    params = ParametresPatrimoine.objects.first()
    validation_active = params.validation_inventaire_active if params else False
    validateurs = params.validateurs_inventaire.all() if params else []
    peut_valider = request.user in validateurs or request.user.is_superuser
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'generer_lignes' and campagne.statut == 'BROUILLON':
            with transaction.atomic():
                campagne.lignes.all().delete()
                biens = Immobilisation.objects.filter(statut__in=['ACTIF', 'EN_PANNE', 'EN_ATTENTE'])
                if campagne.categorie_cible:
                    biens = biens.filter(type_equipement__categorie=campagne.categorie_cible)
                if campagne.batiment_cible:
                    biens = biens.filter(bureau__etage__batiment=campagne.batiment_cible)
                
                lignes_a_creer = [LigneInventairePatrimoine(campagne=campagne, immobilisation=bien) for bien in biens]
                LigneInventairePatrimoine.objects.bulk_create(lignes_a_creer)
                messages.success(request, f"Snapshot généré : {len(lignes_a_creer)} équipements trouvés.")

        elif action == 'demarrer' and campagne.statut == 'BROUILLON':
            if campagne.lignes.count() > 0:
                campagne.statut = 'EN_COURS'
                campagne.save()
                messages.success(request, "L'inventaire a démarré !")
            else:
                messages.error(request, "Générez la liste avant de démarrer.")

        elif action == 'cloturer' and campagne.statut == 'EN_COURS':
            with transaction.atomic():
                lignes_non_pointees = campagne.lignes.filter(etat_constate__isnull=True)
                nb_manquants = lignes_non_pointees.count()
                lignes_non_pointees.update(etat_constate='MANQUANT')

                if validation_active:
                    campagne.statut = 'EN_ATTENTE_VALIDATION'
                    campagne.save()
                    messages.warning(request, f"Inventaire soumis au superviseur. {nb_manquants} machines non scannées ont été déclarées manquantes.")
                else:
                    appliquer_reconciliation_inventaire(request, campagne)
                    messages.success(request, f"Inventaire clôturé. {nb_manquants} manquants enregistrés.")

        elif action == 'valider_inventaire' and campagne.statut == 'EN_ATTENTE_VALIDATION':
            if peut_valider:
                appliquer_reconciliation_inventaire(request, campagne)
                messages.success(request, "✅ Inventaire validé avec succès !")
            else:
                messages.error(request, "⛔ Autorisation refusée : vous n'êtes pas dans la liste des validateurs.")

        return redirect('patrimoine_detail_campagne', campagne_id=campagne.id)

    lignes = campagne.lignes.select_related('immobilisation', 'immobilisation__bureau', 'bureau_constate', 'scanne_par').all()
    total = lignes.count()
    pointes = lignes.exclude(etat_constate__isnull=True).count()
    
    stats = {
        'total': total,
        'pointes': pointes,
        'manquants': lignes.filter(etat_constate='MANQUANT').count(),
        'deplaces': lignes.filter(etat_constate='DEPLACE').count(),
        'rebuts': lignes.filter(etat_constate='A_REFORMER').count(),
        'progression': int((pointes / total * 100)) if total > 0 else 0
    }

    return render(request, 'patrimoine/detail_campagne_inventaire.html', {
        'campagne': campagne,
        'lignes': lignes,
        'stats': stats,
        'validation_active': validation_active,
        'peut_valider': peut_valider
    })

from django.db.models import Count
from django.utils import timezone

def appliquer_reconciliation_inventaire(request, campagne):
    with transaction.atomic():
        a_ete_valide = (campagne.statut == 'EN_ATTENTE_VALIDATION')
        user_actuel = request.user.get_full_name() or request.user.username
        date_heure = timezone.localtime().strftime('%d/%m/%Y à %H:%M')
        
        tech_scan = campagne.lignes.exclude(scanne_par__isnull=True).values(
            'scanne_par__first_name', 'scanne_par__last_name', 'scanne_par__username'
        ).annotate(total=Count('id')).order_by('-total').first()
        
        if tech_scan:
            nom_tech = f"{tech_scan['scanne_par__first_name']} {tech_scan['scanne_par__last_name']}".strip() or tech_scan['scanne_par__username']
        else:
            nom_tech = campagne.cree_par.get_full_name() if campagne.cree_par else "Équipe technique"

        if a_ete_valide:
            phrase_tracabilite = f"Inventorié sur le terrain par : {nom_tech} | Validé par : {user_actuel} le {date_heure}."
        else:
            phrase_tracabilite = f"Inventorié et clôturé par : {user_actuel} le {date_heure} (Sans double validation)."
            
        for ligne in campagne.lignes.filter(etat_constate='DEPLACE'):
            if ligne.bureau_constate:
                immo = ligne.immobilisation
                ancien_bureau = immo.bureau
                immo.bureau = ligne.bureau_constate
                immo.save()
                MouvementPatrimoine.objects.create(
                    immobilisation=immo, type_mouvement='MUTATION',
                    bureau_depart=ancien_bureau, bureau_arrivee=ligne.bureau_constate,
                    motif=f"Déplacé (Inventaire {campagne.reference}). {phrase_tracabilite}", 
                    effectue_par=request.user
                )

        for ligne in campagne.lignes.filter(etat_constate='A_REFORMER'):
            immo = ligne.immobilisation
            immo.statut = 'REFORME'
            immo.bureau = None
            immo.save()
            MouvementPatrimoine.objects.create(
                immobilisation=immo, type_mouvement='REFORME',
                motif=f"Mise au rebut (Inventaire {campagne.reference}). {phrase_tracabilite}", 
                effectue_par=request.user
            )

        campagne.lignes.filter(etat_constate__isnull=True).update(etat_constate='MANQUANT')
        
        for ligne in campagne.lignes.filter(etat_constate='MANQUANT'):
            immo = ligne.immobilisation
            immo.statut = 'DISPARU'  
            immo.save()
            MouvementPatrimoine.objects.create(
                immobilisation=immo, type_mouvement='PERTE',
                motif=f"⚠️ DÉCLARÉ PERDU (Inventaire {campagne.reference}). Non scanné. {phrase_tracabilite}", 
                effectue_par=request.user
            )

        campagne.statut = 'TERMINEE'
        campagne.date_fin_prevue = timezone.localtime().date()
        campagne.save()

# ==============================================================================
# 3. INTERFACE DE SCAN (Mobile / PC)
# ==============================================================================
@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_inventaire')
def audit_scan_inventaire(request, campagne_id):
    campagne = get_object_or_404(CampagneInventairePatrimoine, id=campagne_id, statut='EN_COURS')
    bureaux = Bureau.objects.all().select_related('etage__batiment').order_by('etage__batiment__nom', 'nom')

    if request.method == 'POST':
        code_scanne = request.POST.get('code_patrimoine', '').strip()
        ligne_id = request.POST.get('ligne_id')
        
        if code_scanne and not ligne_id:
            ligne = LigneInventairePatrimoine.objects.filter(campagne=campagne, immobilisation__code_patrimoine__iexact=code_scanne).first()
            if not ligne:
                immo_intruse = Immobilisation.objects.filter(code_patrimoine__iexact=code_scanne).first()
                if immo_intruse:
                    ligne = LigneInventairePatrimoine.objects.create(
                        campagne=campagne, immobilisation=immo_intruse,
                        commentaire="⚠️ Ajouté pendant l'audit (Hors périmètre)"
                    )
                    messages.warning(request, "Attention : Équipement hors périmètre rajouté à l'inventaire.")
                else:
                    messages.error(request, f"⛔ Code inconnu : [{code_scanne}].")
                    return redirect('patrimoine_audit_scan', campagne_id=campagne.id)
            return render(request, 'patrimoine/audit_scan_inventaire.html', {'campagne': campagne, 'ligne_trouvee': ligne, 'bureaux': bureaux})

        if ligne_id:
            ligne = get_object_or_404(LigneInventairePatrimoine, id=ligne_id, campagne=campagne)
            etat = request.POST.get('etat_constate')
            ligne.etat_constate = etat
            ligne.bureau_constate_id = request.POST.get('bureau_constate') if etat == 'DEPLACE' else ligne.immobilisation.bureau_id
            ligne.scanne_par = request.user
            ligne.date_scan = timezone.now()
            ligne.save()
            messages.success(request, f"✅ Pointage enregistré pour {ligne.immobilisation.code_patrimoine}")
            return redirect('patrimoine_audit_scan', campagne_id=campagne.id)

    return render(request, 'patrimoine/audit_scan_inventaire.html', {'campagne': campagne, 'bureaux': bureaux})


from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
import logging

logger = logging.getLogger(__name__)

@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_inventaire')
def imprimer_fiche_comptage(request, campagne_id):
    campagne = get_object_or_404(CampagneInventairePatrimoine, id=campagne_id)
    lignes = campagne.lignes.select_related(
        'immobilisation', 'immobilisation__bureau', 'immobilisation__bureau__etage__batiment'
    ).order_by('immobilisation__bureau__etage__batiment__nom', 'immobilisation__bureau__nom', 'immobilisation__nom_affichage')

    context = {'campagne': campagne, 'lignes': lignes}

    try:
        html_string = render_to_string('patrimoine/imprimer_fiche_comptage.html', context, request=request)
        pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    except Exception as e:
        logger.exception(f"[PDF] Erreur génération fiche de comptage {campagne.reference} : {e}")
        from django.conf import settings
        if settings.DEBUG:
            return HttpResponse(f"<h2>Erreur de génération PDF</h2><p style='color:red;'>{str(e)}</p>", content_type='text/html')
        raise

    response = HttpResponse(pdf_file, content_type='application/pdf')
    nom_fichier = f"Fiche_Comptage_{campagne.reference}.pdf"
    response['Content-Disposition'] = f'inline; filename="{nom_fichier}"'
    return response


@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_rebuts')
def registre_rebuts(request):
    rebuts = Immobilisation.objects.filter(statut='REFORME').select_related(
        'type_equipement', 'service_affectation', 'bureau'
    ).order_by('-id')
    
    for r in rebuts:
        mvt_reforme = r.mouvements.filter(type_mouvement='REFORME').order_by('-date_mouvement').first()
        r.details_tracabilite = mvt_reforme

    return render(request, 'patrimoine/registre_rebuts.html', {'rebuts': rebuts})


@login_required(login_url='/auth/login/')
@patrimoine_required
@verifier_permission('accounts.menu_pat_pertes')
def registre_pertes(request):
    pertes = Immobilisation.objects.filter(statut='DISPARU').select_related(
        'type_equipement', 'service_affectation'
    ).order_by('-id')
    
    for p in pertes:
        mvt_perte = p.mouvements.filter(type_mouvement='PERTE').order_by('-date_mouvement').first()
        p.details_tracabilite = mvt_perte

    return render(request, 'patrimoine/registre_pertes.html', {'pertes': pertes})