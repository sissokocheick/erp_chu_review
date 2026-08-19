# -*- coding: utf-8 -*-
"""Exports/Imports Excel."""
import logging
import json
from decimal import Decimal

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db import transaction, IntegrityError

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from accounts.permissions import verifier_permission

from ..models import (
    Immobilisation, TypeEquipement, CategoriePatrimoine,
    MouvementPatrimoine, ImportPatrimoine, ParametresPatrimoine,
    Batiment, Marque, Modele, Bureau, Etage,
)
from .common import patrimoine_required

logger = logging.getLogger(__name__)


@login_required(login_url='/auth/login/')

@patrimoine_required

@verifier_permission('accounts.menu_pat_registre')

def export_registre_excel(request):

    qs = Immobilisation.objects.select_related(

        'type_equipement__categorie', 'bureau__etage__batiment',

        'service_affectation', 'marque', 'modele',

    ).exclude(statut='EN_ATTENTE').order_by('-date_creation')


    q            = request.GET.get('q', '')

    categorie_id = request.GET.get('categorie', '')

    type_id      = request.GET.get('type', '')

    statut       = request.GET.get('statut', '')

    service_id   = request.GET.get('service', '')

    batiment_id  = request.GET.get('batiment', '')


    if q:

        qs = qs.filter(

            Q(code_patrimoine__icontains=q) | Q(numero_serie__icontains=q) |

            Q(nom_affichage__icontains=q)   | Q(marque__nom__icontains=q)

        ).distinct()

    if categorie_id: qs = qs.filter(type_equipement__categorie_id=categorie_id)

    if type_id:      qs = qs.filter(type_equipement_id=type_id)

    if statut:       qs = qs.filter(statut=statut)

    if service_id:   qs = qs.filter(service_affectation_id=service_id)

    if batiment_id:  qs = qs.filter(bureau__etage__batiment_id=batiment_id)


    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Registre Patrimoine"


    hf = Font(bold=True, color='FFFFFF', size=11)

    hb = PatternFill('solid', fgColor='1C5B96')

    ca = Alignment(horizontal='center', vertical='center')


    headers = [

        'Code patrimoine', 'N° série', 'Désignation', 'Catégorie', 'Type',

        'Marque', 'Modèle', 'Statut', 'Service', 'Bâtiment', 'Bureau',

        'Date acquisition', 'Valeur acquisition (FCFA)', 'VNC (FCFA)', 'Taux amorti (%)',

        'Action requise', 'Créé par', 'Date création', 'Modifié par', 'Date modification',

    ]


    ws.row_dimensions[1].height = 35

    for i, h in enumerate(headers, 1):

        c = ws.cell(row=1, column=i, value=h)

        c.font = hf; c.fill = hb; c.alignment = ca

        ws.column_dimensions[get_column_letter(i)].width = max(16, len(h)+4)


    for row_i, immo in enumerate(qs, 2):

        row = [

            immo.code_patrimoine or '', immo.numero_serie or '', immo.nom_affichage or '',

            immo.type_equipement.categorie.nom if immo.type_equipement_id else '',

            immo.type_equipement.nom if immo.type_equipement_id else '',

            immo.marque.nom if immo.marque_id else '', immo.modele.nom if immo.modele_id else '',

            immo.get_statut_display(),

            immo.service_affectation.nom if immo.service_affectation_id else '',

            str(immo.bureau.batiment.code) if immo.bureau_id else '',

            str(immo.bureau) if immo.bureau_id else '',

            str(immo.date_acquisition) if immo.date_acquisition else '',

            float(immo.valeur_acquisition), float(immo.vnc), float(immo.taux_amorti_pct),

            immo.get_action_requise_display(),

            immo.cree_par.get_full_name() if immo.cree_par_id else '',

            immo.date_creation.strftime('%d/%m/%Y %H:%M') if immo.date_creation else '',

            immo.modifie_par.get_full_name() if immo.modifie_par_id else '',

            immo.date_modification.strftime('%d/%m/%Y %H:%M') if immo.date_modification else '',

        ]

        for col_i, val in enumerate(row, 1):

            ws.cell(row=row_i, column=col_i, value=val)

        if row_i % 2 == 0:

            for col_i in range(1, len(headers)+1):

                ws.cell(row=row_i, column=col_i).fill = PatternFill('solid', fgColor='F0F6FF')


    ws.freeze_panes = 'A2'


    response = HttpResponse(

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    response['Content-Disposition'] = (

        f'attachment; filename="patrimoine_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'

    )

    wb.save(response)

    return response


@login_required(login_url='/auth/login/')

@patrimoine_required

@verifier_permission('accounts.menu_pat_import')

def telecharger_template(request, type_id):

    te = get_object_or_404(TypeEquipement, pk=type_id)


    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = te.nom[:30]


    header_font    = Font(bold=True, color='FFFFFF', size=11)

    header_fill_b  = PatternFill('solid', fgColor='1C5B96')   

    header_fill_p  = PatternFill('solid', fgColor='6F42C1')   

    center         = Alignment(horizontal='center', vertical='center', wrap_text=True)

    thin           = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A1:{get_column_letter(len(COLONNES_FIXES) + len(te.specs_schema))}1')

    ws['A1'] = f"TEMPLATE IMPORT PATRIMOINE — {te.categorie.nom.upper()} / {te.nom.upper()}  |  Colonnes bleues = obligatoires/fixes  |  Colonnes violettes = specs techniques du type"

    ws['A1'].font = Font(bold=True, size=10, color='333333')

    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['A1'].fill = PatternFill('solid', fgColor='F0F3F7')


    ws.row_dimensions[2].height = 40

    for col_idx, nom_col in enumerate(COLONNES_FIXES, start=1):

        cell = ws.cell(row=2, column=col_idx, value=nom_col)

        cell.font = header_font; cell.fill = header_fill_b; cell.alignment = center; cell.border = thin

        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(nom_col) + 4)


    for i, spec in enumerate(te.specs_schema):

        col_idx = len(COLONNES_FIXES) + i + 1

        label = spec.get('label', spec.get('key', f'Spec{i+1}'))

        cell = ws.cell(row=2, column=col_idx, value=label)

        cell.font = header_font; cell.fill = header_fill_p; cell.alignment = center; cell.border = thin

        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(label) + 4)


    ws.row_dimensions[3].height = 20

    exemple = [

        'DIRECTION INFORMATIQUE', 'N', '1er Étage', 'BUREAU INFO', 'HP', 'HP DTP 300 G6 MT', 'SN123456789', 'CHU-INFO-2026-001',

        'UC HP DIRECTION INFO', '2024-01-15', 'BUDGET CHU ANGRÉ', '450000', '2027-01-15', 'RAS', '',

    ]

    for col_idx, val in enumerate(exemple, start=1):

        cell = ws.cell(row=3, column=col_idx, value=val)

        cell.font = Font(italic=True, color='888888', size=10)

        cell.alignment = Alignment(horizontal='left', vertical='center')


    ws.freeze_panes = 'A3'


    ws2 = wb.create_sheet("Guide")

    ws2['A1'] = "GUIDE D'UTILISATION"

    ws2['A1'].font = Font(bold=True, size=14)

    guide = [

        ("", ""), ("Colonne", "Description"),

        ("Service", "Nom exact du service (ex: DIRECTION INFORMATIQUE)"),

        ("Bâtiment", "Code ou nom du bâtiment (ex: N, A, BLOC TECHNIQUE)"),

        ("Étage", "Nom de l'étage (ex: RDC, 1er Étage, Sous-sol)"),

        ("Bureau", "Nom du bureau/salle"),

        ("N° de série", "Numéro de série physique de l'appareil"),

        ("Code patrimoine", "Asset Tag — laisser vide si non encore immatriculé"),

        ("Date acquisition", "Format AAAA-MM-JJ obligatoire pour les calculs d'amortissement"),

        ("Valeur acquisition", "Montant en FCFA — sans espace ni symbole"),

        ("Action requise", f"Valeurs possibles: {', '.join([c[0] for c in Immobilisation.ACTION_CHOICES])}"),

        ("", ""), ("Note", "La ligne 3 de chaque onglet est un exemple — ne pas supprimer, effacer les valeurs si besoin"),

    ]

    for i, (col, desc) in enumerate(guide, start=2):

        ws2.cell(row=i, column=1, value=col).font = Font(bold=True if i == 4 else False)

        ws2.cell(row=i, column=2, value=desc)

    ws2.column_dimensions['A'].width = 25

    ws2.column_dimensions['B'].width = 65


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    response['Content-Disposition'] = f'attachment; filename="template_{te.code}_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    wb.save(response)

    return response


@login_required(login_url='/auth/login/')

@patrimoine_required

@verifier_permission('accounts.menu_pat_import')

def import_excel(request):

    types = TypeEquipement.objects.filter(est_actif=True).select_related('categorie')


    if request.method == 'POST':

        fichier   = request.FILES.get('fichier')

        type_id   = request.POST.get('type_equipement')


        if not fichier or not type_id:

            messages.error(request, "Veuillez choisir un type et un fichier.")

            return render(request, 'patrimoine/import.html', {'types': types})


        if not fichier.name.endswith(('.xlsx', '.xls')):

            messages.error(request, "Format accepté : .xlsx ou .xls uniquement.")

            return render(request, 'patrimoine/import.html', {'types': types})


        te = get_object_or_404(TypeEquipement, pk=type_id)


        try:

            wb = openpyxl.load_workbook(fichier, data_only=True)

            ws = wb.active


            header_row  = None

            header_map  = {}

            for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), start=1):

                row_vals = [str(c).strip() if c else '' for c in row]

                if any('service' in v.lower() for v in row_vals):

                    header_row = row_idx

                    for col_idx, val in enumerate(row_vals):

                        if val:

                            header_map[val.lower().strip()] = col_idx

                    break


            if header_row is None:

                messages.error(request, "En-têtes introuvables. Utilisez le template fourni.")

                return redirect('patrimoine_import')


            def get_val(row, key_fragments):

                for map_key, col_idx in header_map.items():

                    if any(frag.lower() in map_key for frag in key_fragments):

                        val = row[col_idx] if col_idx < len(row) else None

                        return str(val).strip() if val else ''

                return ''


            nb_crees = nb_maj = nb_err = 0

            log_err  = []


            for row_num, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):

                if not any(row): continue

                if row_num == header_row + 1 and get_val(row, ['service']).upper() in ('SERVICE', 'DIRECTION INFORMATIQUE', 'NOM DU SERVICE'):

                    continue


                try:

                    from core.models import Service

                    from stock.models import Fournisseur


                    nom_bat = get_val(row, ['timent', 'bat'])

                    nom_eta = get_val(row, ['tage'])

                    nom_bur = get_val(row, ['bureau', 'salle'])

                    nom_svc = get_val(row, ['service'])


                    batiment = Batiment.objects.filter(Q(code__iexact=nom_bat) | Q(nom__iexact=nom_bat)).first()

                    if not batiment and nom_bat: batiment = Batiment.objects.create(code=nom_bat[:10].upper(), nom=nom_bat.upper(), cree_par=request.user)


                    etage = None

                    if batiment and nom_eta: etage, _ = Etage.objects.get_or_create(batiment=batiment, nom=nom_eta.upper(), defaults={'cree_par': request.user})


                    bureau = None

                    if etage and nom_bur: bureau, _ = Bureau.objects.get_or_create(etage=etage, nom=nom_bur.upper(), defaults={'cree_par': request.user})


                    service = Service.objects.filter(nom__iexact=nom_svc).first() if nom_svc else None


                    nom_marque = get_val(row, ['marque'])

                    nom_modele = get_val(row, ['mod'])

                    marque = Marque.objects.get_or_create(nom=nom_marque.upper(), defaults={'cree_par': request.user})[0] if nom_marque else None

                    modele = Modele.objects.get_or_create(marque=marque, nom=nom_modele.upper(), defaults={'cree_par': request.user})[0] if marque and nom_modele else None


                    code_pat  = get_val(row, ['asset', 'code', 'inventaire'])

                    num_serie = get_val(row, ['rie', 'sn'])

                    nom_aff   = get_val(row, ['nom', 'affich', 'equipement'])

                    date_acq_raw = get_val(row, ['acquisition', 'date'])

                    valeur_raw   = get_val(row, ['valeur', 'montant', 'fcfa'])

                    garantie_raw = get_val(row, ['garantie', 'expiration'])


                    date_acq = None

                    if date_acq_raw:

                        from datetime import datetime

                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):

                            try: date_acq = datetime.strptime(str(date_acq_raw)[:10], fmt).date(); break

                            except ValueError: continue


                    valeur = Decimal('0.00')

                    if valeur_raw:

                        try: valeur = Decimal(str(valeur_raw).replace(' ', '').replace(',', '.'))

                        except Exception:

                            logger.debug("[import_excel] Valeur '%s' non convertible en Decimal", valeur_raw)


                    garantie = None

                    if garantie_raw:

                        from datetime import datetime

                        try: garantie = datetime.strptime(str(garantie_raw)[:10], '%Y-%m-%d').date()

                        except (ValueError, TypeError) as e:
                            logger.debug("[import_excel] Parsing garantie echoue : %s", e)


                    action = get_val(row, ['action'])

                    if action not in [c[0] for c in Immobilisation.ACTION_CHOICES]: action = 'RAS'

                    notes  = get_val(row, ['note'])


                    specs = {}

                    for spec in te.specs_schema:

                        key = spec['key']; label = spec.get('label', key).lower()

                        val = get_val(row, [label, key])

                        if val: specs[key] = val


                    nom_fourn = get_val(row, ['fournisseur'])

                    fournisseur = Fournisseur.objects.filter(raison_sociale__icontains=nom_fourn).first() if nom_fourn else None


                    lookup = {}

                    if code_pat and code_pat.upper() not in ('NA', 'N/A', ''): lookup['code_patrimoine'] = code_pat

                    elif num_serie: lookup['numero_serie'] = num_serie

                    else: lookup = None


                    defaults = {

                        'type_equipement': te, 'nom_affichage': nom_aff or '', 'numero_serie': num_serie or '',

                        'marque': marque, 'modele': modele, 'bureau': bureau, 'service_affectation': service,

                        'date_acquisition': date_acq, 'valeur_acquisition': valeur, 'garantie_expiration': garantie,

                        'action_requise': action, 'notes': notes, 'specs_techniques': specs, 'fournisseur': fournisseur,

                        'statut': 'EN_ATTENTE' if not code_pat or code_pat.upper() in ('NA', 'N/A') else 'ACTIF',

                        'cree_par': request.user, 'reference_inventaire': fichier.name[:50],

                    }


                    if lookup:

                        obj, created = Immobilisation.objects.update_or_create(**lookup, defaults=defaults)

                        if created: nb_crees += 1

                        else: nb_maj += 1

                    else:

                        Immobilisation.objects.create(**defaults)

                        nb_crees += 1


                except Exception as e:

                    nb_err += 1

                    log_err.append({'ligne': row_num, 'erreur': str(e)})


            statut_log = 'OK' if nb_err == 0 else ('PARTIEL' if nb_crees + nb_maj > 0 else 'ECHEC')

            log = ImportPatrimoine.objects.create(

                type_equipement=te, nb_lignes_traitees=nb_crees + nb_maj + nb_err, nb_crees=nb_crees,

                nb_mis_a_jour=nb_maj, nb_erreurs=nb_err, log_erreurs=log_err, statut=statut_log, cree_par=request.user,

            )


            if nb_err == 0: messages.success(request, f"✅ Import réussi — {nb_crees} créés, {nb_maj} mis à jour.")

            else: messages.warning(request, f"⚠️ Import partiel — {nb_crees} créés, {nb_maj} mis à jour, {nb_err} erreurs.")


            return redirect('patrimoine_import_log', pk=log.pk)


        except Exception as e:

            messages.error(request, f"❌ Erreur lecture fichier : {e}")

            return redirect('patrimoine_import')


    context = {'types': types, 'logs_recents': ImportPatrimoine.objects.order_by('-date_creation')[:5]}

    return render(request, 'patrimoine/import.html', context)


@login_required(login_url='/auth/login/')

@patrimoine_required

@verifier_permission('accounts.menu_pat_import')

def detail_import_log(request, pk):

    log = get_object_or_404(ImportPatrimoine, pk=pk)

    return render(request, 'patrimoine/import_log.html', {'log': log})
